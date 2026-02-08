"""Discover and parse historical evaluation runs from local files."""

import csv
import json
import os
import logging
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional
import time

# Configure logger for run discovery
logger = logging.getLogger(__name__)

# Increase CSV field size limit to handle large cell values (e.g., long LLM outputs)
# Default is 128KB which is too small for verbose model outputs
# Set to 500MB - generous but not unlimited to prevent memory issues
CSV_FIELD_SIZE_LIMIT = 500 * 1024 * 1024  # 500MB
csv.field_size_limit(CSV_FIELD_SIZE_LIMIT)

DEFAULT_RESULTS_DIR = "qym_results"

# Error score constant - errors are always scored as 0
# This is the Python equivalent of metrics.js getRowScore()
ERROR_SCORE = 0.0


def parse_metric_score(value: Any) -> Optional[float]:
    """Parse a metric score from common boolean/numeric formats."""
    if value is None:
        return None
    raw = str(value).strip()
    if raw == "":
        return None
    lowered = raw.lower()
    if lowered in {"n/a", "na", "none"}:
        return None
    if raw in {"✓"} or lowered in {"true", "yes", "y"}:
        return 1.0
    if raw in {"✗"} or lowered in {"false", "no", "n"}:
        return 0.0
    if lowered in {"1", "1.0"}:
        return 1.0
    if lowered in {"0", "0.0"}:
        return 0.0
    if raw.endswith("%"):
        try:
            return float(raw[:-1].strip()) / 100.0
        except (ValueError, TypeError):
            return None
    try:
        return float(raw)
    except (ValueError, TypeError):
        return None


def normalize_metric_score(value: Any) -> Optional[str]:
    """Normalize a metric score to a numeric string for CSV storage."""
    if value is None:
        return ""
    raw = str(value).strip()
    if raw == "":
        return ""
    parsed = parse_metric_score(raw)
    if parsed is None:
        return None
    return f"{parsed:g}"


def parse_run_metadata(raw_metadata: Any) -> Dict[str, Any]:
    """Parse run_metadata payload into a dict."""
    if isinstance(raw_metadata, dict):
        return raw_metadata
    if raw_metadata in (None, ""):
        return {}
    try:
        parsed = json.loads(str(raw_metadata))
        return parsed if isinstance(parsed, dict) else {}
    except (json.JSONDecodeError, TypeError, ValueError):
        return {}


def merge_run_metadata(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Merge run_metadata from rows, preferring the latest non-empty values."""
    merged: Dict[str, Any] = {}
    for row in rows:
        metadata = parse_run_metadata(row.get("run_metadata", "{}"))
        if not metadata:
            continue
        for key, value in metadata.items():
            if value in (None, ""):
                continue
            merged[key] = value
    return merged


def parse_total_items(value: Any) -> Optional[int]:
    """Parse a positive integer total_items value from metadata."""
    if value in (None, ""):
        return None
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        try:
            parsed = int(float(str(value)))
        except (TypeError, ValueError):
            return None
    return parsed if parsed > 0 else None


def is_error_row(row: Dict[str, Any], metrics: List[str]) -> bool:
    """Check if a row represents an error.

    This is the SINGLE SOURCE OF TRUTH for error detection in Python.
    Frontend equivalent: metrics.js isErrorRow()
    """
    output = str(row.get("output", "") or "")
    if output.startswith("ERROR:"):
        return True
    # Also check if first metric score contains ERROR
    if metrics:
        score_str = str(row.get(f"{metrics[0]}_score", "") or "")
        if "ERROR" in score_str or score_str == "N/A":
            return True
    return False


@dataclass
class RunInfo:
    """Metadata for a single evaluation run."""
    run_id: str
    task_name: str
    model_name: str
    dataset_name: str
    timestamp: datetime
    file_path: str
    metrics: List[str]
    total_items: int
    success_count: int
    error_count: int
    metric_averages: Dict[str, float] = field(default_factory=dict)
    avg_latency_ms: float = 0.0
    langfuse_url: Optional[str] = None
    langfuse_dataset_id: Optional[str] = None
    langfuse_run_id: Optional[str] = None

    @property
    def success_rate(self) -> float:
        if self.total_items == 0:
            return 0.0
        return self.success_count / self.total_items

    def to_dict(self) -> Dict[str, Any]:
        return {
            "run_id": self.run_id,
            "task_name": self.task_name,
            "model_name": self.model_name,
            "dataset_name": self.dataset_name,
            "timestamp": self.timestamp.isoformat(),
            "file_path": self.file_path,
            "metrics": self.metrics,
            "metric_averages": self.metric_averages,
            "total_items": self.total_items,
            "success_count": self.success_count,
            "error_count": self.error_count,
            "success_rate": self.success_rate,
            "avg_latency_ms": self.avg_latency_ms,
            "langfuse_url": self.langfuse_url,
            "langfuse_dataset_id": self.langfuse_dataset_id,
            "langfuse_run_id": self.langfuse_run_id,
        }


@dataclass
class RunIndex:
    """Hierarchical index: task -> model -> runs."""
    tasks: Dict[str, Dict[str, List[RunInfo]]] = field(default_factory=dict)
    last_updated: datetime = field(default_factory=datetime.now)

    def to_dict(self) -> Dict[str, Any]:
        result = {"tasks": {}, "last_updated": self.last_updated.isoformat()}
        for task_name, models in self.tasks.items():
            result["tasks"][task_name] = {}
            for model_name, runs in models.items():
                result["tasks"][task_name][model_name] = [r.to_dict() for r in runs]
        return result


def strip_model_provider(model_name: str) -> str:
    """Strip provider prefix from model name (e.g., 'qwen/qwen3-coder' -> 'qwen3-coder')."""
    if not model_name:
        return model_name
    parts = model_name.split("/")
    return parts[-1] if len(parts) > 1 else model_name


class RunDiscovery:
    """Scan results directory and build index of historical runs."""

    def __init__(self, results_dir: str = DEFAULT_RESULTS_DIR):
        self.results_dir = Path(results_dir)
        self._cache: Optional[RunIndex] = None
        self._cache_time: Optional[float] = None
        self._cache_ttl = 30  # seconds

    def scan(self, force_refresh: bool = False) -> RunIndex:
        """Scan results directory and return hierarchical index."""
        # Check cache
        if not force_refresh and self._cache is not None:
            if self._cache_time and (time.time() - self._cache_time) < self._cache_ttl:
                return self._cache

        index = RunIndex()

        if not self.results_dir.exists():
            self._cache = index
            self._cache_time = time.time()
            return index

        # Traverse: results_dir/{task}/{model}/{date}/*.csv
        for task_dir in self.results_dir.iterdir():
            if not task_dir.is_dir():
                continue
            task_name = task_dir.name

            for model_dir in task_dir.iterdir():
                if not model_dir.is_dir():
                    continue
                model_name = model_dir.name

                for date_dir in model_dir.iterdir():
                    if not date_dir.is_dir():
                        continue

                    # Support both .csv and .xlsx files
                    for result_file in list(date_dir.glob("*.csv")) + list(date_dir.glob("*.xlsx")):
                        run_info = self._parse_result_file(result_file, task_name, model_name)
                        if run_info:
                            if task_name not in index.tasks:
                                index.tasks[task_name] = {}
                            if model_name not in index.tasks[task_name]:
                                index.tasks[task_name][model_name] = []
                            index.tasks[task_name][model_name].append(run_info)

        # Sort runs by timestamp (newest first)
        for task_name in index.tasks:
            for model_name in index.tasks[task_name]:
                index.tasks[task_name][model_name].sort(
                    key=lambda r: r.timestamp, reverse=True
                )

        index.last_updated = datetime.now()
        self._cache = index
        self._cache_time = time.time()
        return index

    def _parse_result_file(
        self, file_path: Path, task_name: str, model_name: str
    ) -> Optional[RunInfo]:
        """Parse result file (CSV or XLSX) to extract run metadata."""
        if file_path.suffix.lower() == ".xlsx":
            return self._parse_xlsx_file(file_path, task_name, model_name)
        return self._parse_csv_file(file_path, task_name, model_name)

    def _parse_csv_file(
        self, file_path: Path, task_name: str, model_name: str
    ) -> Optional[RunInfo]:
        """Parse CSV file to extract run metadata."""
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                fieldnames = reader.fieldnames or []

                # Extract metrics from column names ending with _score
                # Exclude metadata columns (containing __meta__)
                metrics = [
                    col.replace("_score", "")
                    for col in fieldnames
                    if col.endswith("_score") and "__meta__" not in col
                ]

                # Read all rows to count totals
                rows = list(reader)
                processed_items = len(rows)

                if processed_items == 0:
                    return None

                # Get run info from first row
                first_row = rows[0]
                run_name = first_row.get("run_name", "")
                dataset_name = first_row.get("dataset_name", "unknown")

                # Merge run_metadata across rows, because incremental checkpoints
                # may only include Langfuse IDs after the first few items finish.
                metadata = merge_run_metadata(rows)
                actual_model = strip_model_provider(
                    str(metadata.get("model") or model_name)
                )
                langfuse_url = metadata.get("langfuse_url") or None
                langfuse_dataset_id = metadata.get("langfuse_dataset_id") or None
                langfuse_run_id = metadata.get("langfuse_run_id") or None
                declared_total_items = parse_total_items(metadata.get("total_items"))
                total_items = max(processed_items, declared_total_items or 0)

                # Count successes vs errors and calculate metric averages
                error_count = 0
                metric_sums: Dict[str, float] = {m: 0.0 for m in metrics}
                metric_counts: Dict[str, int] = {m: 0 for m in metrics}
                latency_sum = 0.0
                latency_count = 0

                for row in rows:
                    row_is_error = is_error_row(row, metrics)
                    if row_is_error:
                        error_count += 1

                    # Accumulate metric scores (errors = 0)
                    for m in metrics:
                        score_str = row.get(f"{m}_score", "")
                        if row_is_error:
                            # Errors are scored as 0
                            metric_sums[m] += ERROR_SCORE
                            metric_counts[m] += 1
                        else:
                            score = parse_metric_score(score_str)
                            if score is not None:
                                metric_sums[m] += score
                                metric_counts[m] += 1

                    # Accumulate latency (time column is in seconds)
                    time_str = row.get("time", "")
                    try:
                        latency_sec = float(time_str)
                        latency_sum += latency_sec * 1000  # Convert to ms
                        latency_count += 1
                    except (ValueError, TypeError):
                        pass

                success_count = processed_items - error_count

                # Calculate metric averages
                metric_averages = {}
                for m in metrics:
                    if metric_counts[m] > 0:
                        metric_averages[m] = metric_sums[m] / metric_counts[m]
                    else:
                        metric_averages[m] = 0.0

                # Calculate average latency
                avg_latency_ms = latency_sum / latency_count if latency_count > 0 else 0.0

                # Parse timestamp from run_name
                timestamp = self._extract_timestamp(run_name)

                return RunInfo(
                    run_id=run_name or file_path.stem,
                    task_name=task_name,
                    model_name=actual_model,
                    dataset_name=dataset_name,
                    timestamp=timestamp,
                    file_path=str(file_path.resolve()),  # Use absolute path
                    metrics=metrics,
                    total_items=total_items,
                    success_count=success_count,
                    error_count=error_count,
                    metric_averages=metric_averages,
                    avg_latency_ms=avg_latency_ms,
                    langfuse_url=langfuse_url,
                    langfuse_dataset_id=langfuse_dataset_id,
                    langfuse_run_id=langfuse_run_id,
                )
        except csv.Error as e:
            error_msg = str(e)
            if "field larger than field limit" in error_msg:
                logger.warning(
                    f"[RunDiscovery] Skipping '{file_path.name}': CSV field exceeds {CSV_FIELD_SIZE_LIMIT // (1024*1024)}MB limit. "
                    f"File may contain extremely large cell values."
                )
            else:
                logger.warning(f"[RunDiscovery] Skipping '{file_path.name}': CSV parsing error - {error_msg}")
            return None
        except UnicodeDecodeError as e:
            logger.warning(f"[RunDiscovery] Skipping '{file_path.name}': File encoding error - {e}")
            return None
        except Exception as e:
            logger.warning(f"[RunDiscovery] Skipping '{file_path.name}': Unexpected error - {type(e).__name__}: {e}")
            return None

    def _parse_xlsx_file(
        self, file_path: Path, task_name: str, model_name: str
    ) -> Optional[RunInfo]:
        """Parse XLSX file to extract run metadata."""
        try:
            import openpyxl
        except ImportError:
            return None

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active

            # Get header from first row
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            fieldnames = [str(cell) if cell else "" for cell in header_row]

            # Read all data rows
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {fieldnames[i]: row[i] for i in range(len(fieldnames)) if i < len(row)}
                rows.append(row_dict)

            wb.close()

            if not rows:
                return None

            # Extract metrics from column names ending with _score
            # Exclude metadata columns (containing __meta__)
            metrics = [
                col.replace("_score", "")
                for col in fieldnames
                if col.endswith("_score") and "__meta__" not in col
            ]

            processed_items = len(rows)
            first_row = rows[0]
            run_name = str(first_row.get("run_name", "") or "")
            dataset_name = str(first_row.get("dataset_name", "unknown") or "unknown")

            metadata = merge_run_metadata(rows)
            actual_model = strip_model_provider(
                str(metadata.get("model") or model_name)
            )
            langfuse_url = metadata.get("langfuse_url") or None
            langfuse_dataset_id = metadata.get("langfuse_dataset_id") or None
            langfuse_run_id = metadata.get("langfuse_run_id") or None
            declared_total_items = parse_total_items(metadata.get("total_items"))
            total_items = max(processed_items, declared_total_items or 0)

            # Count successes vs errors and calculate metric averages
            error_count = 0
            metric_sums: Dict[str, float] = {m: 0.0 for m in metrics}
            metric_counts: Dict[str, int] = {m: 0 for m in metrics}
            latency_sum = 0.0
            latency_count = 0

            for row in rows:
                row_is_error = is_error_row(row, metrics)
                if row_is_error:
                    error_count += 1

                # Accumulate metric scores (errors = 0)
                for m in metrics:
                    score_val = row.get(f"{m}_score")
                    if row_is_error:
                        # Errors are scored as 0
                        metric_sums[m] += ERROR_SCORE
                        metric_counts[m] += 1
                    else:
                        try:
                            score = float(score_val) if score_val is not None else None
                            if score is not None:
                                metric_sums[m] += score
                                metric_counts[m] += 1
                        except (ValueError, TypeError):
                            pass

                # Accumulate latency
                time_val = row.get("time")
                try:
                    latency_sec = float(time_val) if time_val is not None else None
                    if latency_sec is not None:
                        latency_sum += latency_sec * 1000
                        latency_count += 1
                except (ValueError, TypeError):
                    pass

            success_count = processed_items - error_count

            # Calculate metric averages
            metric_averages = {}
            for m in metrics:
                if metric_counts[m] > 0:
                    metric_averages[m] = metric_sums[m] / metric_counts[m]
                else:
                    metric_averages[m] = 0.0

            avg_latency_ms = latency_sum / latency_count if latency_count > 0 else 0.0
            timestamp = self._extract_timestamp(run_name)

            return RunInfo(
                run_id=run_name or file_path.stem,
                task_name=task_name,
                model_name=actual_model,
                dataset_name=dataset_name,
                timestamp=timestamp,
                file_path=str(file_path),
                metrics=metrics,
                total_items=total_items,
                success_count=success_count,
                error_count=error_count,
                metric_averages=metric_averages,
                avg_latency_ms=avg_latency_ms,
                langfuse_url=langfuse_url,
                langfuse_dataset_id=langfuse_dataset_id,
                langfuse_run_id=langfuse_run_id,
            )
        except Exception as e:
            logger.warning(f"[RunDiscovery] Skipping '{file_path.name}': XLSX parsing error - {type(e).__name__}: {e}")
            return None

    def _extract_timestamp(self, run_name: str) -> datetime:
        """Extract timestamp from run_name pattern: task-model-YYMMDD-HHMM[-counter]."""
        match = re.search(r"-(\d{6}-\d{4})(?:-\d+)?$", run_name or "")
        if match:
            try:
                return datetime.strptime(match.group(1), "%y%m%d-%H%M")
            except ValueError:
                pass
        return datetime.now()

    def get_run_data(self, file_path: str) -> Dict[str, Any]:
        """Load full run data and transform to UI snapshot format."""
        path = Path(file_path)
        if not path.is_absolute() and file_path and os.path.exists("/" + file_path):
            path = Path("/" + file_path)
        
        # Try multiple path resolutions:
        # 1. As-is (could be absolute or relative to CWD)
        # 2. Relative to results_dir parent (if file_path includes results_dir name)
        # 3. As a subpath of results_dir
        if not path.exists():
            # Try resolving relative to results_dir's parent
            if self.results_dir.exists():
                # If file_path is like "qym_results/task/model/date/file.csv"
                # and results_dir is "qym_results", try parent/file_path
                alt_path = self.results_dir.parent / file_path
                if alt_path.exists():
                    path = alt_path
                else:
                    # Try as subpath: results_dir / (file_path minus results_dir name)
                    results_name = self.results_dir.name
                    if file_path.startswith(results_name + "/"):
                        sub_path = file_path[len(results_name) + 1:]
                        alt_path = self.results_dir / sub_path
                        if alt_path.exists():
                            path = alt_path
        
        if not path.exists():
            return {"error": f"File not found: {file_path}"}

        # Handle xlsx files
        if path.suffix.lower() == ".xlsx":
            return self._get_xlsx_run_data(path)

        try:
            with open(path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                fieldnames = reader.fieldnames or []
                rows = list(reader)
        except csv.Error as e:
            error_msg = str(e)
            if "field larger than field limit" in error_msg:
                return {
                    "error": f"CSV field exceeds {CSV_FIELD_SIZE_LIMIT // (1024*1024)}MB size limit. "
                    f"This file contains extremely large cell values that cannot be loaded."
                }
            return {"error": f"CSV parsing error: {error_msg}"}
        except UnicodeDecodeError as e:
            return {"error": f"File encoding error: {e}. Try re-saving the file as UTF-8."}
        except Exception as e:
            return {"error": f"{type(e).__name__}: {e}"}

        if not rows:
            return {"error": "Empty file - no data rows found"}

        # Extract metrics (exclude metadata columns containing __meta__)
        metric_names = [
            col.replace("_score", "")
            for col in fieldnames
            if col.endswith("_score") and "__meta__" not in col
        ]

        # Build run info
        first_row = rows[0]
        run_name = first_row.get("run_name", "")
        dataset_name = first_row.get("dataset_name", "")

        # Parse config
        config = {}
        try:
            config = json.loads(first_row.get("run_config", "{}"))
        except (json.JSONDecodeError, TypeError):
            pass

        # Parse metadata
        metadata = {}
        try:
            metadata = json.loads(first_row.get("run_metadata", "{}"))
        except (json.JSONDecodeError, TypeError):
            pass

        # Transform rows to UI snapshot format
        ui_rows = []
        stats = {"total": len(rows), "completed": 0, "in_progress": 0, "pending": 0, "failed": 0}

        for idx, row in enumerate(rows):
            output = row.get("output", "")
            is_error = output.startswith("ERROR:") or output.startswith("ERROR ")

            # Determine status
            status = "completed"
            if is_error:
                status = "error"
                stats["failed"] += 1
            else:
                stats["completed"] += 1

            # Parse time
            time_val = row.get("time", "0")
            try:
                latency_ms = float(time_val) * 1000
            except (ValueError, TypeError):
                latency_ms = 0

            # Extract metric values
            metric_values = []
            metric_meta = {}
            for m in metric_names:
                score = row.get(f"{m}_score", "")
                metric_values.append(score)
                # Collect additional metric fields (metadata)
                meta = {}
                for col in fieldnames:
                    # New format: {metric}__meta__{field}
                    if col.startswith(f"{m}__meta__"):
                        field_name = col[len(f"{m}__meta__"):]
                        if field_name == "json":
                            raw = row.get(col, "")
                            if raw:
                                try:
                                    parsed = json.loads(raw)
                                    if isinstance(parsed, dict):
                                        meta.update({k: str(v) for k, v in parsed.items()})
                                except Exception:
                                    meta[field_name] = raw
                        else:
                            meta[field_name] = row.get(col, "")
                    # Legacy format: {metric}_{field} (but not {metric}_score)
                    elif col.startswith(f"{m}_") and col != f"{m}_score" and "__meta__" not in col:
                        field_name = col[len(f"{m}_"):]
                        # Only use legacy format if it's not another metric's score column
                        if not field_name.endswith("_score"):
                            meta[field_name] = row.get(col, "")
                if meta:
                    metric_meta[m] = meta

            # Build UI row
            input_full = row.get("input", "")
            output_full = row.get("output", "")
            expected_full = row.get("expected_output", "")

            # Optional: per-item task start timestamp (epoch ms)
            task_started_at_ms = None
            raw_started = row.get("task_started_at_ms", None)
            if raw_started not in (None, ""):
                try:
                    task_started_at_ms = int(float(raw_started))
                except (ValueError, TypeError):
                    task_started_at_ms = None

            ui_row = {
                "index": idx,
                "item_id": row.get("item_id", str(idx)),  # Use item_id from CSV, fallback to index
                "status": status,
                "input": input_full,
                "input_full": input_full,
                "output": output_full,
                "output_full": output_full,
                "expected": expected_full,
                "expected_full": expected_full,
                "time": time_val,
                "latency_ms": latency_ms,
                "task_started_at_ms": task_started_at_ms,
                "trace_id": row.get("trace_id", ""),
                "trace_url": "",  # Historical runs don't have live trace URLs
                "metric_values": metric_values,
                "metric_meta": metric_meta,
            }
            ui_rows.append(ui_row)

        # Calculate success rate
        stats["success_rate"] = (stats["completed"] / stats["total"] * 100) if stats["total"] > 0 else 0

        return {
            "run": {
                "dataset_name": dataset_name,
                "run_name": run_name,
                "file_path": str(path.resolve()),
                "metric_names": metric_names,
                "config": config,
                "metadata": metadata,
                "langfuse_host": "",
                "langfuse_project_id": "",
            },
            "snapshot": {
                "rows": ui_rows,
                "stats": stats,
                "metric_names": metric_names,
            },
        }

    def _get_xlsx_run_data(self, path: Path) -> Dict[str, Any]:
        """Load xlsx run data and transform to UI snapshot format."""
        try:
            import openpyxl
        except ImportError:
            return {"error": "openpyxl package is required to read xlsx files. Install with: pip install openpyxl"}

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active

            # Get header from first row
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            fieldnames = [str(cell) if cell else "" for cell in header_row]

            # Read all data rows
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {fieldnames[i]: row[i] for i in range(len(fieldnames)) if i < len(row)}
                rows.append(row_dict)

            wb.close()
        except Exception as e:
            return {"error": f"XLSX parsing error: {type(e).__name__}: {e}"}

        if not rows:
            return {"error": "Empty file - no data rows found"}

        # Extract metrics (exclude metadata columns containing __meta__)
        metric_names = [
            col.replace("_score", "")
            for col in fieldnames
            if col.endswith("_score") and "__meta__" not in col
        ]

        # Build run info
        first_row = rows[0]
        run_name = str(first_row.get("run_name", "") or "")
        dataset_name = str(first_row.get("dataset_name", "") or "")

        # Parse config
        config = {}
        try:
            config_str = first_row.get("run_config", "{}")
            if config_str:
                config = json.loads(str(config_str))
        except (json.JSONDecodeError, TypeError):
            pass

        # Parse metadata
        metadata = {}
        try:
            metadata_str = first_row.get("run_metadata", "{}")
            if metadata_str:
                metadata = json.loads(str(metadata_str))
        except (json.JSONDecodeError, TypeError):
            pass

        # Transform rows to UI snapshot format
        ui_rows = []
        stats = {"total": len(rows), "completed": 0, "in_progress": 0, "pending": 0, "failed": 0}

        for idx, row in enumerate(rows):
            output = str(row.get("output", "") or "")
            is_error = output.startswith("ERROR:") or output.startswith("ERROR ")

            # Determine status
            status = "completed"
            if is_error:
                status = "error"
                stats["failed"] += 1
            else:
                stats["completed"] += 1

            # Parse time
            time_val = row.get("time", 0)
            try:
                latency_ms = float(time_val) * 1000 if time_val else 0
            except (ValueError, TypeError):
                latency_ms = 0

            # Extract metric values
            metric_values = []
            metric_meta = {}
            for m in metric_names:
                score = row.get(f"{m}_score", "")
                metric_values.append(score if score is not None else "")
                # Collect additional metric fields (metadata)
                meta = {}
                for col in fieldnames:
                    # New format: {metric}__meta__{field}
                    if col.startswith(f"{m}__meta__"):
                        field_name = col[len(f"{m}__meta__"):]
                        if field_name == "json":
                            raw = row.get(col, "")
                            if raw:
                                try:
                                    parsed = json.loads(str(raw))
                                    if isinstance(parsed, dict):
                                        meta.update({k: str(v) for k, v in parsed.items()})
                                except Exception:
                                    meta[field_name] = str(raw)
                        else:
                            meta[field_name] = row.get(col, "")
                    # Legacy format: {metric}_{field} (but not {metric}_score)
                    elif col.startswith(f"{m}_") and col != f"{m}_score" and "__meta__" not in col:
                        field_name = col[len(f"{m}_"):]
                        # Only use legacy format if it's not another metric's score column
                        if not field_name.endswith("_score"):
                            meta[field_name] = row.get(col, "")
                if meta:
                    metric_meta[m] = meta

            # Build UI row
            input_full = str(row.get("input", "") or "")
            output_full = str(row.get("output", "") or "")
            expected_full = str(row.get("expected_output", "") or "")

            # Optional: per-item task start timestamp (epoch ms)
            task_started_at_ms = None
            raw_started = row.get("task_started_at_ms", None)
            if raw_started not in (None, ""):
                try:
                    task_started_at_ms = int(float(raw_started))
                except (ValueError, TypeError):
                    task_started_at_ms = None

            ui_row = {
                "index": idx,
                "item_id": str(row.get("item_id", "") or "") or str(idx),  # Use item_id from xlsx, fallback to index
                "status": status,
                "input": input_full,
                "input_full": input_full,
                "output": output_full,
                "output_full": output_full,
                "expected": expected_full,
                "expected_full": expected_full,
                "time": str(time_val) if time_val else "0",
                "latency_ms": latency_ms,
                "task_started_at_ms": task_started_at_ms,
                "trace_id": str(row.get("trace_id", "") or ""),
                "trace_url": "",
                "metric_values": metric_values,
                "metric_meta": metric_meta,
            }
            ui_rows.append(ui_row)

        # Calculate success rate
        stats["success_rate"] = (stats["completed"] / stats["total"] * 100) if stats["total"] > 0 else 0

        return {
            "run": {
                "dataset_name": dataset_name,
                "run_name": run_name,
                "file_path": str(path.resolve()),
                "metric_names": metric_names,
                "config": config,
                "metadata": metadata,
                "langfuse_host": "",
                "langfuse_project_id": "",
            },
            "snapshot": {
                "rows": ui_rows,
                "stats": stats,
                "metric_names": metric_names,
            },
        }

    def update_metric_score(
        self, file_path: str, row_index: int, metric_name: str, new_score: Any
    ) -> Dict[str, Any]:
        """Update a metric score in a CSV run file and record audit metadata."""
        path = Path(file_path)

        # Try multiple path resolutions (same as get_run_data)
        if not path.exists():
            if self.results_dir.exists():
                alt_path = self.results_dir.parent / file_path
                if alt_path.exists():
                    path = alt_path
                else:
                    results_name = self.results_dir.name
                    if file_path.startswith(results_name + "/"):
                        sub_path = file_path[len(results_name) + 1 :]
                        alt_path = self.results_dir / sub_path
                        if alt_path.exists():
                            path = alt_path

        if not path.exists():
            return {"error": f"File not found: {file_path}"}

        # Security: ensure the resolved path is within results_dir
        try:
            results_abs = os.path.abspath(self.results_dir)
            resolved = os.path.abspath(path)
            if resolved != results_abs and not resolved.startswith(results_abs + os.sep):
                return {"error": "Access denied"}
        except Exception:
            return {"error": "Access denied"}

        if path.suffix.lower() == ".xlsx":
            return {"error": "XLSX runs are read-only"}

        try:
            with open(path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                fieldnames = reader.fieldnames or []
                rows = list(reader)
        except csv.Error as e:
            error_msg = str(e)
            if "field larger than field limit" in error_msg:
                return {
                    "error": f"CSV field exceeds {CSV_FIELD_SIZE_LIMIT // (1024*1024)}MB size limit."
                }
            return {"error": f"CSV parsing error: {error_msg}"}
        except UnicodeDecodeError as e:
            return {"error": f"File encoding error: {e}. Try re-saving the file as UTF-8."}
        except Exception as e:
            return {"error": f"{type(e).__name__}: {e}"}

        if row_index < 0 or row_index >= len(rows):
            return {"error": f"Row index out of range: {row_index}"}

        score_col = f"{metric_name}_score"
        if score_col not in fieldnames:
            return {"error": f"Metric not found: {metric_name}"}

        meta_modified = f"{metric_name}__meta__modified"
        meta_original = f"{metric_name}__meta__original_score"
        if meta_modified not in fieldnames:
            fieldnames.append(meta_modified)
        if meta_original not in fieldnames:
            fieldnames.append(meta_original)

        # Normalize new score to string for CSV
        new_score_str = normalize_metric_score(new_score)
        if new_score_str is None:
            return {"error": "Invalid score value"}
        row = rows[row_index]
        current_score = row.get(score_col, "")
        original_score = row.get(meta_original, "")
        if original_score in (None, ""):
            original_score = current_score

        row[score_col] = new_score_str
        row[meta_original] = original_score
        original_comp = parse_metric_score(original_score)
        if original_comp is None and str(original_score).strip() != "":
            original_comp = str(original_score).strip()
        new_comp = parse_metric_score(new_score_str)
        if new_comp is None and str(new_score_str).strip() != "":
            new_comp = str(new_score_str).strip()
        row[meta_modified] = "true" if new_comp != original_comp else "false"

        # Ensure all rows have the audit columns
        for r in rows:
            if meta_modified not in r:
                r[meta_modified] = ""
            if meta_original not in r:
                r[meta_original] = ""

        try:
            with open(path, "w", encoding="utf-8", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
                writer.writeheader()
                writer.writerows(rows)
        except Exception as e:
            return {"error": f"Failed to write CSV: {type(e).__name__}: {e}"}

        return {"ok": True}
