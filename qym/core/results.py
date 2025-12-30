"""Evaluation results container and analysis."""

import csv
import json
import os
import statistics
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence
import re

from rich import box
from rich.align import Align
from rich.columns import Columns
from rich.console import Console, Group
from rich.panel import Panel
from rich.rule import Rule
from rich.table import Table
from rich.text import Text


console = Console()


class EvaluationResult:
    """Container for evaluation results with analysis capabilities."""
    
    def __init__(self, dataset_name: str, run_name: str, metrics: List[str], run_metadata: Optional[Dict[str, Any]] = None, run_config: Optional[Dict[str, Any]] = None):
        """
        Initialize results container.
        
        Args:
            dataset_name: Name of the evaluated dataset
            run_name: Name of this evaluation run
            metrics: List of metric names used
        """
        self.dataset_name = dataset_name
        self.run_name = run_name
        self.metrics = metrics
        self.run_metadata = run_metadata or {}
        self.run_config = run_config or {}
        self.start_time = datetime.now()
        self.end_time = None
        self.last_saved_path: Optional[str] = None
        self._save_notice_consumed = False
        self.langfuse_url: Optional[str] = None  # URL to Langfuse dataset run

        # Results storage
        self.inputs = {}  # item_id -> input data
        self.metadatas = {}  # item_id -> metadata dict
        self.results = {}  # item_id -> result dict
        self.errors = {}   # item_id -> {"error": str, "trace_id": Optional[str]}

    def add_input(self, item_id: str, task_input: Any):
        """Add input data for an item."""
        self.inputs[item_id] = task_input

    def add_metadata(self, item_id: str, metadata: Dict[str, Any]):
        """Add metadata for an item."""
        self.metadatas[item_id] = metadata

    def add_result(self, item_id: str, result: Dict[str, Any]):
        """Add a successful evaluation result."""
        self.results[item_id] = result

    def add_error(self, item_id: str, error: str, trace_id: Optional[str] = None):
        """Add an evaluation error."""
        self.errors[item_id] = {"error": error, "trace_id": trace_id}
    
    def finish(self):
        """Mark evaluation as finished."""
        self.end_time = datetime.now()
    
    @property
    def total_items(self) -> int:
        """Total number of evaluated items."""
        return len(self.results) + len(self.errors)
    
    @property
    def success_rate(self) -> float:
        """Percentage of successful evaluations."""
        if self.total_items == 0:
            return 0.0
        return len(self.results) / self.total_items
    
    @property
    def duration(self) -> Optional[float]:
        """Evaluation duration in seconds."""
        if self.end_time:
            return (self.end_time - self.start_time).total_seconds()
        return None
    
    def get_metric_stats(self, metric_name: str) -> Dict[str, float]:
        """
        Get statistics for a specific metric.
        
        Returns dict with: mean, std, min, max, success_rate
        """
        scores = []
        errors = 0
        
        for result in self.results.values():
            if 'scores' in result and metric_name in result['scores']:
                score = result['scores'][metric_name]
                if isinstance(score, dict):
                    if 'error' in score:
                        errors += 1
                        continue
                    if 'score' in score and isinstance(score['score'], (int, float)):
                        scores.append(float(score['score']))
                        continue
                if isinstance(score, (int, float)):
                    scores.append(float(score))
                elif isinstance(score, bool):
                    scores.append(1.0 if score else 0.0)
        
        if not scores:
            return {
                'mean': 0.0,
                'std': 0.0,
                'min': 0.0,
                'max': 0.0,
                'success_rate': 0.0
            }
        
        return {
            'mean': statistics.mean(scores),
            'std': statistics.stdev(scores) if len(scores) > 1 else 0.0,
            'min': min(scores),
            'max': max(scores),
            'success_rate': len(scores) / (len(scores) + errors)
        }
    
    def get_timing_stats(self) -> Dict[str, float]:
        """
        Get timing statistics for all evaluations.
        
        Returns dict with: mean, std, min, max, total
        """
        times = []
        
        for result in self.results.values():
            if 'time' in result and isinstance(result['time'], (int, float)):
                times.append(float(result['time']))
        
        if not times:
            return {
                'mean': 0.0,
                'std': 0.0,
                'min': 0.0,
                'max': 0.0,
                'total': 0.0
            }
        
        return {
            'mean': statistics.mean(times),
            'std': statistics.stdev(times) if len(times) > 1 else 0.0,
            'min': min(times),
            'max': max(times),
            'total': sum(times)
        }
    
    def summary(self) -> str:
        """Generate a text summary of results."""
        lines = []
        lines.append(f"Evaluation Results: {self.run_name}")
        lines.append(f"Dataset: {self.dataset_name}")
        lines.append(f"Total Items: {self.total_items}")
        lines.append(f"Success Rate: {self.success_rate:.1%}")
        
        if self.duration:
            lines.append(f"Duration: {self.duration:.1f}s")
        
        lines.append("\nMetric Results:")
        for metric in self.metrics:
            stats = self.get_metric_stats(metric)
            lines.append(f"  {metric}:")
            lines.append(f"    Mean: {stats['mean']:.3f}")
            lines.append(f"    Std:  {stats['std']:.3f}")
            lines.append(f"    Range: [{stats['min']:.3f}, {stats['max']:.3f}]")
        
        if self.errors:
            lines.append(f"\nErrors: {len(self.errors)} items failed")
            # Show all errors
            for item_id, error_info in self.errors.items():
                error_msg = error_info["error"] if isinstance(error_info, dict) else error_info
                lines.append(f"  - {item_id}: {error_msg}")
        
        return "\n".join(lines)
    
    def consume_saved_notice(self, include_run_name: bool = False) -> Optional[str]:
        """Return a formatted saved-results notice once per save."""
        if not self.last_saved_path or self._save_notice_consumed:
            return None
        self._save_notice_consumed = True
        if include_run_name and self.run_name:
            return f"{self.run_name}: {self.last_saved_path}"
        return self.last_saved_path
    
    
    def print_summary(self, html_url: Optional[str] = None, *, force: bool = False):
        """Print a consolidated summary that matches the multi-run view."""
        if not force and not summary_display_enabled():
            return
        panel = render_results_summary([self])
        console.print()
        console.print(panel)
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert results to dictionary format."""
        metric_stats = {
            metric: self.get_metric_stats(metric) 
            for metric in self.metrics
        }
        
        return {
            'dataset_name': self.dataset_name,
            'run_name': self.run_name,
            'start_time': self.start_time.isoformat(),
            'end_time': self.end_time.isoformat() if self.end_time else None,
            'duration': self.duration,
            'total_items': self.total_items,
            'success_rate': self.success_rate,
            'metrics': self.metrics,
            'metric_stats': metric_stats,
            'langfuse_url': self.langfuse_url,
            'inputs': self.inputs,
            'metadatas': self.metadatas,
            'results': self.results,
            'errors': self.errors
        }
    
    def failed_items(self) -> List[str]:
        """Get list of failed item IDs."""
        return list(self.errors.keys())
    
    def successful_items(self) -> List[str]:
        """Get list of successful item IDs."""
        return list(self.results.keys())
    
    def save_json(self, filepath: Optional[str] = None, output_dir: str = ".") -> str:
        """
        Save results to JSON file.
        
        Args:
            filepath: Optional custom filepath. If not provided, generates one.
            output_dir: Directory to save to if filepath is not provided.
            
        Returns:
            Path to the saved file
        """
        if filepath is None:
            filepath = self._default_save_path("json", output_dir)
        
        filepath = Path(filepath)
        filepath.parent.mkdir(parents=True, exist_ok=True)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(self.to_dict(), f, indent=2, default=str, ensure_ascii=False)
        
        return str(filepath)
    
    def save_csv(self, filepath: Optional[str] = None, output_dir: str = ".") -> str:
        """
        Save results to CSV file for spreadsheet analysis.
        
        Args:
            filepath: Optional custom filepath. If not provided, generates one.
            output_dir: Directory to save to if filepath is not provided.
            
        Returns:
            Path to the saved file
        """
        if filepath is None:
            filepath = self._default_save_path("csv", output_dir)
        
        filepath = Path(filepath)
        filepath.parent.mkdir(parents=True, exist_ok=True)
        
        # Build metric metadata field list across results for consistent columns
        def flatten_meta(md: Dict[str, Any]) -> Dict[str, Any]:
            flat: Dict[str, Any] = {}
            try:
                for k, v in (md or {}).items():
                    if isinstance(v, dict):
                        for k2, v2 in v.items():
                            flat[f"{k}_{k2}"] = v2
                    else:
                        flat[str(k)] = v
            except Exception:
                pass
            return flat

        meta_fields: Dict[str, List[str]] = {m: [] for m in self.metrics}
        for result in self.results.values():
            scores = result.get('scores', {})
            for m in self.metrics:
                sc = scores.get(m)
                if isinstance(sc, dict):
                    md = sc.get('metadata') if 'metadata' in sc else {}
                    fl = flatten_meta(md if isinstance(md, dict) else {})
                    for key in fl.keys():
                        if key not in meta_fields[m]:
                            meta_fields[m].append(key)

        # Prepare header
        base_fields = [
            'dataset_name', 'run_name', 'run_metadata', 'run_config',
            'trace_id', 'item_id', 'input', 'item_metadata', 'output', 'expected_output', 'time'
        ]
        metric_fields: List[str] = []
        for m in self.metrics:
            metric_fields.append(f'{m}_score')
            for k in meta_fields[m]:
                # Use __meta__ separator to distinguish from score columns
                metric_fields.append(f'{m}__meta__{k}')
        header = base_fields + metric_fields

        # Prepare row formatter helpers
        def main_score(val: Any) -> Any:
            if isinstance(val, dict):
                if 'error' in val:
                    return f"ERROR: {val['error']}"
                if 'score' in val:
                    return val.get('score')
            return val

        # Assemble rows
        rows: List[Dict[str, Any]] = []
        for item_id, result in self.results.items():
            # Get item metadata
            item_metadata = self.metadatas.get(item_id, '')
            if isinstance(item_metadata, dict):
                item_metadata = json.dumps(item_metadata, ensure_ascii=False)

            row: Dict[str, Any] = {
                'dataset_name': self.dataset_name,
                'run_name': self.run_name,
                'run_metadata': json.dumps(self.run_metadata, ensure_ascii=False),
                'run_config': json.dumps(self.run_config, ensure_ascii=False),
                'trace_id': result.get('trace_id', ''),
                'item_id': item_id,
                'input': result.get('input', ''),
                'item_metadata': str(item_metadata),
                'output': result.get('output', ''),
                'expected_output': result.get('expected', ''),
                'time': result.get('time', 0.0),
            }
            scores = result.get('scores', {})
            for m in self.metrics:
                val = main_score(scores.get(m))
                row[f'{m}_score'] = val
                md: Dict[str, Any] = {}
                sc = scores.get(m)
                if isinstance(sc, dict) and 'metadata' in sc and isinstance(sc['metadata'], dict):
                    md = flatten_meta(sc['metadata'])
                for k in meta_fields[m]:
                    row[f'{m}__meta__{k}'] = md.get(k, '')
            rows.append(row)

        # Add failed items as rows too
        for item_id, error_info in self.errors.items():
            # Handle both old format (string) and new format (dict with error and trace_id)
            if isinstance(error_info, dict):
                error_msg = error_info.get("error", str(error_info))
                error_trace_id = error_info.get("trace_id", "")
            else:
                error_msg = str(error_info)
                error_trace_id = ""

            # Get input and metadata for failed items
            task_input = self.inputs.get(item_id, '')
            if isinstance(task_input, dict):
                task_input = json.dumps(task_input, ensure_ascii=False)

            item_metadata = self.metadatas.get(item_id, '')
            if isinstance(item_metadata, dict):
                item_metadata = json.dumps(item_metadata, ensure_ascii=False)

            row = {
                'dataset_name': self.dataset_name,
                'run_name': self.run_name,
                'run_metadata': json.dumps(self.run_metadata, ensure_ascii=False),
                'run_config': json.dumps(self.run_config, ensure_ascii=False),
                'trace_id': error_trace_id or '',
                'item_id': item_id,
                'input': str(task_input),
                'item_metadata': str(item_metadata),
                'output': f'ERROR: {error_msg}',
                'expected_output': '',
                'time': 0.0,
            }
            for m in self.metrics:
                row[f'{m}_score'] = 'N/A'
                for k in meta_fields[m]:
                    row[f'{m}__meta__{k}'] = ''
            rows.append(row)

        # Write CSV
        if rows:
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=header)
                writer.writeheader()
                writer.writerows(rows)

        return str(filepath)

    def save_excel(self, filepath: Optional[str] = None, output_dir: str = ".") -> str:
        """
        Save results to xlsx file for spreadsheet analysis.

        Args:
            filepath: Optional custom filepath. If not provided, generates one.
            output_dir: Directory to save to if filepath is not provided.

        Returns:
            Path to the saved file
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. Install it with: pip install openpyxl"
            )

        if filepath is None:
            filepath = self._default_save_path("xlsx", output_dir)

        filepath = Path(filepath)
        filepath.parent.mkdir(parents=True, exist_ok=True)

        # Build metric metadata field list across results for consistent columns
        def flatten_meta(md: Dict[str, Any]) -> Dict[str, Any]:
            flat: Dict[str, Any] = {}
            try:
                for k, v in (md or {}).items():
                    if isinstance(v, dict):
                        for k2, v2 in v.items():
                            flat[f"{k}_{k2}"] = v2
                    else:
                        flat[str(k)] = v
            except Exception:
                pass
            return flat

        meta_fields: Dict[str, List[str]] = {m: [] for m in self.metrics}
        for result in self.results.values():
            scores = result.get('scores', {})
            for m in self.metrics:
                sc = scores.get(m)
                if isinstance(sc, dict):
                    md = sc.get('metadata') if 'metadata' in sc else {}
                    fl = flatten_meta(md if isinstance(md, dict) else {})
                    for key in fl.keys():
                        if key not in meta_fields[m]:
                            meta_fields[m].append(key)

        # Prepare header (same as CSV)
        base_fields = [
            'dataset_name', 'run_name', 'run_metadata', 'run_config',
            'trace_id', 'item_id', 'input', 'item_metadata', 'output', 'expected_output', 'time'
        ]
        metric_fields: List[str] = []
        for m in self.metrics:
            metric_fields.append(f'{m}_score')
            for k in meta_fields[m]:
                # Use __meta__ separator to distinguish from score columns
                metric_fields.append(f'{m}__meta__{k}')
        header = base_fields + metric_fields

        # Prepare row formatter helpers
        def main_score(val: Any) -> Any:
            if isinstance(val, dict):
                if 'error' in val:
                    return f"ERROR: {val['error']}"
                if 'score' in val:
                    return val.get('score')
            return val

        # Assemble rows (same logic as CSV)
        rows: List[Dict[str, Any]] = []
        for item_id, result in self.results.items():
            item_metadata = self.metadatas.get(item_id, '')
            if isinstance(item_metadata, dict):
                item_metadata = json.dumps(item_metadata, ensure_ascii=False)

            row: Dict[str, Any] = {
                'dataset_name': self.dataset_name,
                'run_name': self.run_name,
                'run_metadata': json.dumps(self.run_metadata, ensure_ascii=False),
                'run_config': json.dumps(self.run_config, ensure_ascii=False),
                'trace_id': result.get('trace_id', ''),
                'item_id': item_id,
                'input': result.get('input', ''),
                'item_metadata': str(item_metadata),
                'output': result.get('output', ''),
                'expected_output': result.get('expected', ''),
                'time': result.get('time', 0.0),
            }
            scores = result.get('scores', {})
            for m in self.metrics:
                val = main_score(scores.get(m))
                row[f'{m}_score'] = val
                md: Dict[str, Any] = {}
                sc = scores.get(m)
                if isinstance(sc, dict) and 'metadata' in sc and isinstance(sc['metadata'], dict):
                    md = flatten_meta(sc['metadata'])
                for k in meta_fields[m]:
                    row[f'{m}__meta__{k}'] = md.get(k, '')
            rows.append(row)

        # Add failed items as rows too
        for item_id, error_info in self.errors.items():
            # Handle both old format (string) and new format (dict with error and trace_id)
            if isinstance(error_info, dict):
                error_msg = error_info.get("error", str(error_info))
                error_trace_id = error_info.get("trace_id", "")
            else:
                error_msg = str(error_info)
                error_trace_id = ""

            task_input = self.inputs.get(item_id, '')
            if isinstance(task_input, dict):
                task_input = json.dumps(task_input, ensure_ascii=False)

            item_metadata = self.metadatas.get(item_id, '')
            if isinstance(item_metadata, dict):
                item_metadata = json.dumps(item_metadata, ensure_ascii=False)

            row = {
                'dataset_name': self.dataset_name,
                'run_name': self.run_name,
                'run_metadata': json.dumps(self.run_metadata, ensure_ascii=False),
                'run_config': json.dumps(self.run_config, ensure_ascii=False),
                'trace_id': error_trace_id or '',
                'item_id': item_id,
                'input': str(task_input),
                'item_metadata': str(item_metadata),
                'output': f'ERROR: {error_msg}',
                'expected_output': '',
                'time': 0.0,
            }
            for m in self.metrics:
                row[f'{m}_score'] = 'N/A'
                for k in meta_fields[m]:
                    row[f'{m}__meta__{k}'] = ''
            rows.append(row)

        # Write Excel using openpyxl directly
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"

        # Write header
        for col_idx, field in enumerate(header, start=1):
            ws.cell(row=1, column=col_idx, value=field)

        # Write data rows
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, field in enumerate(header, start=1):
                value = row_data.get(field, '')
                # Convert non-string types to string for complex objects
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, ensure_ascii=False)
                ws.cell(row=row_idx, column=col_idx, value=value)

        wb.save(filepath)

        return str(filepath)

    def _default_save_path(self, extension: str, output_dir: str) -> str:
        """Create default save path with hierarchy: qym_results/task/model/date/filename."""
        ts = _extract_run_timestamp(self.run_name) or datetime.now()
        timestamp_str = ts.strftime("%y%m%d-%H%M")
        date_dir = ts.strftime("%Y-%m-%d")

        model_name = (
            (self.run_metadata.get("model") if isinstance(self.run_metadata, dict) else None)
            or (self.run_config.get("model") if isinstance(self.run_config, dict) else None)
            or "nomodel"
        )
        task_name = _task_from_run(self.run_name, model_name) or "task"

        task_safe = _sanitize_path_component(task_name)
        model_safe = _sanitize_path_component(str(model_name))

        # Extract counter suffix if present after timestamp (e.g., -YYMMDD-HHMM-1)
        # Must follow the timestamp pattern to avoid matching the timestamp itself
        counter_match = re.search(r"-\d{6}-\d{4}-(\d+)$", self.run_name or "")
        counter_suffix = f"-{counter_match.group(1)}" if counter_match else ""

        # Format: [base]-[dataset]-[model]-[timestamp][-counter].csv
        filename = f"{task_safe}-{self.dataset_name}-{model_safe}-{timestamp_str}{counter_suffix}.{extension}"

        return str(
            Path(output_dir)
            / task_safe
            / model_safe
            / date_dir
            / filename
        )
    
    def save(self, format: str = "json", filepath: Optional[str] = None, output_dir: str = ".") -> str:
        """
        Save results in specified format.

        Args:
            format: Export format - "json", "csv", or "xlsx"
            filepath: Optional custom filepath
            output_dir: Directory to save to if filepath is not provided.

        Returns:
            Path to the saved file
        """
        saved_path: Optional[str] = None
        if format.lower() == "json":
            saved_path = self.save_json(filepath, output_dir=output_dir)
        elif format.lower() == "csv":
            saved_path = self.save_csv(filepath, output_dir=output_dir)
        elif format.lower() == "xlsx":
            saved_path = self.save_excel(filepath, output_dir=output_dir)
        else:
            raise ValueError(f"Unsupported format: {format}. Use 'json', 'csv', or 'xlsx'.")
        self.last_saved_path = saved_path
        self._save_notice_consumed = False
        return saved_path


def render_results_summary(
    results: Sequence[EvaluationResult],
    title: Optional[str] = None,
) -> Panel:
    """Render a shared Rich panel for one or many EvaluationResult objects."""
    run_count = len(results)
    default_title = "[bold cyan]Multi-Run Summary[/bold cyan]" if run_count > 1 else "[bold cyan]Run Summary[/bold cyan]"
    panel_title = title or default_title

    # Overview stats
    total_items = sum(r.total_items for r in results)
    total_success = sum(len(r.results) for r in results)
    overview_ratio = (total_success / total_items * 100.0) if total_items else 0.0

    header_text = Text(f"{run_count} run{'s' if run_count != 1 else ''} • {total_items} items • {overview_ratio:.1f}% success")
    header_text.stylize("dim")

    summary_table = _build_run_summary_table(results)
    metric_section = _build_metric_section(results)
    error_panel = _build_error_panel(results)

    content: List[Any] = [
        Align.center(Text(panel_title.replace("[bold cyan]", "").replace("[/bold cyan]", ""), style="bold cyan")),
        Align.center(header_text),
        Rule(style="dim"),
        summary_table,
    ]

    if metric_section is not None:
        content.extend(
            [
                Rule(style="dim"),
                Align.center(Text("Metric Performance", style="bold")),
                metric_section,
            ]
        )

    if error_panel is not None:
        content.extend([Rule(style="dim"), error_panel])

    return Panel(
        Group(*content),
        title=panel_title,
        border_style="cyan",
        padding=(1, 2),
        expand=True,
    )


def _build_run_summary_table(results: Sequence[EvaluationResult]) -> Table:
    table = Table(
        box=box.SIMPLE_HEAD,
        expand=True,
        show_lines=False,
        padding=(0, 1),
        header_style="bold",
    )
    table.add_column("Run", style="cyan", overflow="fold", ratio=2)
    table.add_column("Dataset", overflow="fold", ratio=2)
    table.add_column("Items", justify="right", width=8)
    table.add_column("Success", justify="right", width=10)
    table.add_column("Avg Latency", justify="right", width=12)
    table.add_column("Duration", justify="right", width=12)
    table.add_column("Errors", justify="right", width=8)

    if not results:
        table.add_row("-", "-", "-", "-", "-", "-", "-")
        return table

    for result in results:
        timing_stats = result.get_timing_stats()
        avg_latency = timing_stats.get("mean") or 0.0
        run_label = _label_with_model(result.run_name, result.run_metadata or {})
        dataset_label = Text(result.dataset_name)
        table.add_row(
            run_label,
            dataset_label,
            str(result.total_items),
            f"{result.success_rate * 100:.1f}%",
            f"{avg_latency:.2f}s",
            _human_duration(result.duration),
            str(len(result.errors)),
        )
    return table


def _build_metric_section(results: Sequence[EvaluationResult]):
    if not results:
        return None

    metric_panels = []
    for result in results:
        metric_table = Table(
            box=box.SIMPLE,
            show_header=True,
            expand=True,
            header_style="bold",
            padding=(0, 1),
        )
        metric_table.add_column("Metric", style="cyan", ratio=2)
        metric_table.add_column("Mean", justify="right", width=10)
        metric_table.add_column("Std", justify="right", width=10)
        metric_table.add_column("Min", justify="right", width=10)
        metric_table.add_column("Max", justify="right", width=10)
        metric_table.add_column("Success", justify="right", width=10)

        if not result.metrics:
            metric_table.add_row("-", "-", "-", "-", "-", "-")
        else:
            for metric in result.metrics:
                stats = result.get_metric_stats(metric)
                metric_table.add_row(
                    metric,
                    f"{stats['mean']:.3f}",
                    f"{stats['std']:.3f}",
                    f"{stats['min']:.3f}",
                    f"{stats['max']:.3f}",
                    f"{stats['success_rate'] * 100:.1f}%",
                )

        panel_title = f"{result.run_name} Metrics" if len(results) > 1 else "Metric Details"
        metric_panels.append(
            Panel(
                metric_table,
                title=panel_title,
                border_style="cyan",
                padding=(0, 1),
            )
        )

    if len(metric_panels) == 1:
        return metric_panels[0]
    return Columns(metric_panels, expand=True, equal=False)


def _build_error_panel(results: Sequence[EvaluationResult]) -> Optional[Panel]:
    error_lines: List[str] = []
    for result in results:
        if not result.errors:
            continue
        error_lines.append(f"{result.run_name}: {len(result.errors)} failures")
        for item_id, message in result.errors.items():
            error_lines.append(f"  • {item_id}: {message}")

    if not error_lines:
        return None

    return Panel(
        "\n".join(error_lines),
        title="[bold red]Errors[/bold red]",
        border_style="red",
        padding=(0, 1),
    )


def _human_duration(seconds: Optional[float]) -> str:
    if seconds is None:
        return "--"
    total_seconds = max(0, int(seconds))
    if total_seconds < 60:
        return f"{total_seconds}s"
    minutes, sec = divmod(total_seconds, 60)
    if minutes < 60:
        return f"{minutes}m {sec:02d}s"
    hours, minutes = divmod(minutes, 60)
    return f"{hours}h {minutes:02d}m"


def _label_with_model(name: str, metadata: Dict[str, Any]) -> Text:
    label = Text(name, style="bold")
    model = metadata.get("model")
    if model:
        label.append(f"\n{model}", style="dim")
    return label


def summary_display_enabled() -> bool:
    """Return True when summaries should render (opt-in via env flag)."""
    value = os.environ.get("QYM_SHOW_SUMMARY", "")
    return value.lower() in {"1", "true", "yes", "on"}


_RUN_ID_RE = re.compile(r"^(?P<base>.+)-(?P<model>.+)-(?P<ts>\d{6}-\d{4})$")


def _strip_run_suffix(name: str) -> str:
    """Strip timestamp/model suffix to recover base run name."""
    if not name:
        return ""
    match = _RUN_ID_RE.match(name)
    if not match:
        return name
    return match.group("base")


def _task_from_run(run_name: str, model_name: Optional[str]) -> str:
    """Derive task name by removing counter/timestamp/model suffixes.

    Run name format: {task}-{model}-{YYMMDD-HHMM}[-{counter}]
    Example: ai_assistant-gpt-4o-mini-251126-1712-3
    Should return: ai_assistant
    """
    base = run_name or ""

    # 1. Strip timestamp with optional counter suffix in one pass
    # Pattern: -YYMMDD-HHMM or -YYMMDD-HHMM-N (where N is counter)
    base = re.sub(r"-\d{6}-\d{4}(?:-\d+)?$", "", base)

    # 2. Strip model if known
    if model_name and base.endswith(f"-{model_name}"):
        base = base[: -len(model_name) - 1]

    return base


def _extract_run_timestamp(run_name: str) -> Optional[datetime]:
    """Extract timestamp from run name, accounting for optional counter suffix.

    Run name format: {task}-{model}-{YYMMDD-HHMM}[-{counter}]
    """
    name = run_name or ""
    # Match timestamp with optional counter suffix
    match = re.search(r"-(?P<ts>\d{6}-\d{4})(?:-\d+)?$", name)
    if not match:
        return None
    ts_str = match.group("ts")
    try:
        return datetime.strptime(ts_str, "%y%m%d-%H%M")
    except Exception:
        return None


def _sanitize_path_component(value: str) -> str:
    """Basic filesystem-safe component."""
    if not value:
        return "unknown"
    cleaned = value.replace(os.sep, "_").replace("\\", "_")
    return cleaned.strip() or "unknown"
