"""Excel chart embedding functionality for LLM evaluation results."""

import io
import tempfile
from typing import Dict, List, Any, Optional, Tuple, TYPE_CHECKING
from pathlib import Path
import statistics

if TYPE_CHECKING:
    from llm_eval.core.results import EvaluationResult

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.chart import (
        BarChart, LineChart, PieChart, ScatterChart, 
        Reference, Series
    )
    from openpyxl.chart.axis import DateAxis, ValuesAxis
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import ColorScaleRule
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

# EvaluationResult will be passed as parameter to avoid circular imports


class ExcelChartExporter:
    """
    Exports LLM evaluation results to Excel with embedded charts.
    
    Creates professional Excel reports with:
    - Data tables with proper formatting
    - Embedded charts (bar, line, pie, scatter)
    - Summary statistics
    - Professional styling for executive reports
    """
    
    # Professional styling constants
    if OPENPYXL_AVAILABLE:
        HEADER_STYLE = {
            'font': Font(bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center')
        }
        
        DATA_STYLE = {
            'font': Font(color='343A40'),
            'alignment': Alignment(horizontal='left', vertical='center')
        }
        
        NUMERIC_STYLE = {
            'font': Font(color='343A40'),
            'alignment': Alignment(horizontal='right', vertical='center'),
            'number_format': '0.000'
        }
        
        SUCCESS_STYLE = {
            'fill': PatternFill(start_color='A8E6CF', end_color='A8E6CF', fill_type='solid')
        }
        
        FAILURE_STYLE = {
            'fill': PatternFill(start_color='FFD3B6', end_color='FFD3B6', fill_type='solid')
        }
    else:
        HEADER_STYLE = {}
        DATA_STYLE = {}
        NUMERIC_STYLE = {}
        SUCCESS_STYLE = {}
        FAILURE_STYLE = {}
    
    def __init__(self):
        """Initialize Excel chart exporter."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel export functionality. "
                "Install with: pip install openpyxl"
            )
        
        self.workbook = None
        self.color_palette = [
            '2E86AB', 'A23B72', 'F18F01', 'C73E1D', 
            '6C757D', '17A2B8', '28A745', 'FFC107'
        ]
    
    def create_excel_report(
        self, 
        results,  # EvaluationResult object
        filename: str,
        include_raw_data: bool = True,
        include_charts: bool = True
    ) -> str:
        """
        Create comprehensive Excel report with charts.
        
        Args:
            results: EvaluationResult object
            filename: Output Excel filename
            include_raw_data: Whether to include raw evaluation data
            include_charts: Whether to include charts
            
        Returns:
            Path to created Excel file
        """
        self.workbook = Workbook()
        
        # Remove default sheet
        self.workbook.remove(self.workbook.active)
        
        # Create sheets
        self._create_summary_sheet(results)
        
        if include_raw_data:
            self._create_detailed_results_sheet(results)
        
        if include_charts:
            self._create_charts_sheet(results)
        
        # Save workbook
        filepath = Path(filename)
        filepath.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(filepath)
        
        return str(filepath)
    
    def _create_summary_sheet(self, results) -> None:  # EvaluationResult object
        """Create summary sheet with key metrics and statistics."""
        ws = self.workbook.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = f"Evaluation Report: {results.run_name}"
        ws['A1'].font = Font(size=16, bold=True, color='2E86AB')
        ws.merge_cells('A1:E1')
        
        # Basic information
        row = 3
        info_data = [
            ("Dataset:", results.dataset_name),
            ("Run Name:", results.run_name),
            ("Total Items:", results.total_items),
            ("Successful Items:", len(results.results)),
            ("Failed Items:", len(results.errors)),
            ("Success Rate:", f"{results.success_rate:.1%}"),
            ("Duration:", f"{results.duration:.1f}s" if results.duration else "N/A")
        ]
        
        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Metric statistics
        row += 2
        ws[f'A{row}'] = "Metric Statistics"
        ws[f'A{row}'].font = Font(size=14, bold=True, color='2E86AB')
        row += 1
        
        # Headers
        headers = ['Metric', 'Mean', 'Std Dev', 'Min', 'Max', 'Success Rate']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.HEADER_STYLE['font']
            cell.fill = self.HEADER_STYLE['fill']
            cell.alignment = self.HEADER_STYLE['alignment']
        row += 1
        
        # Metric data
        for metric in results.metrics:
            stats = results.get_metric_stats(metric)
            data = [
                metric,
                stats['mean'],
                stats['std'],
                stats['min'],
                stats['max'],
                stats['success_rate']
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if col == 1:  # Metric name
                    cell.font = self.DATA_STYLE['font']
                    cell.alignment = self.DATA_STYLE['alignment']
                else:  # Numeric values
                    cell.font = self.NUMERIC_STYLE['font']
                    cell.alignment = self.NUMERIC_STYLE['alignment']
                    if isinstance(value, float) and col != 6:  # Not success rate
                        cell.number_format = '0.000'
                    elif col == 6:  # Success rate
                        cell.number_format = '0.0%'
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add simple bar chart for metrics
        if len(results.metrics) > 0:
            self._add_metric_summary_chart(ws, results, row + 2)
    
    def _create_detailed_results_sheet(self, results) -> None:  # EvaluationResult object
        """Create detailed results sheet with all evaluation data."""
        ws = self.workbook.create_sheet("Detailed Results")
        
        # Headers
        headers = ['Item ID', 'Output', 'Success', 'Time (s)']
        headers.extend([f'{metric}' for metric in results.metrics])
        headers.append('Errors')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_STYLE['font']
            cell.fill = self.HEADER_STYLE['fill']
            cell.alignment = self.HEADER_STYLE['alignment']
        
        # Data rows
        row = 2
        
        # Successful results
        for item_id, result in results.results.items():
            data = [
                item_id,
                str(result.get('output', ''))[:100] + ('...' if len(str(result.get('output', ''))) > 100 else ''),
                'Yes',
                result.get('time', 0.0)
            ]
            
            # Add metric scores
            if 'scores' in result:
                for metric in results.metrics:
                    score = result['scores'].get(metric, 'N/A')
                    if isinstance(score, dict) and 'error' in score:
                        data.append('ERROR')
                    else:
                        data.append(score)
            else:
                data.extend(['N/A'] * len(results.metrics))
            
            data.append('')  # No errors for successful items
            
            # Write data
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if col in [4]:  # Time column
                    cell.number_format = '0.000'
                elif col > 4 and col <= 4 + len(results.metrics):  # Metric columns
                    if isinstance(value, (int, float)):
                        cell.number_format = '0.000'
            row += 1
        
        # Failed results
        for item_id, error in results.errors.items():
            data = [
                item_id,
                f'ERROR: {str(error)[:100]}',
                'No',
                0.0
            ]
            data.extend(['N/A'] * len(results.metrics))
            data.append(str(error))
            
            # Write data
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if col == 3:  # Success column
                    cell.font = Font(color='C73E1D')  # Red for failed
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add conditional formatting for success column
        success_col = 3
        success_range = f"{ws.cell(row=2, column=success_col).coordinate}:{ws.cell(row=row-1, column=success_col).coordinate}"
        
        # Color successful items green, failed items red
        from openpyxl.formatting.rule import CellIsRule
        green_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
        red_fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
        
        ws.conditional_formatting.add(success_range, 
            CellIsRule(operator='equal', formula=['"Yes"'], fill=green_fill))
        ws.conditional_formatting.add(success_range, 
            CellIsRule(operator='equal', formula=['"No"'], fill=red_fill))
    
    def _create_charts_sheet(self, results) -> None:  # EvaluationResult object
        """Create charts sheet with embedded visualizations."""
        ws = self.workbook.create_sheet("Charts")
        
        # Title
        ws['A1'] = "Evaluation Charts"
        ws['A1'].font = Font(size=16, bold=True, color='2E86AB')
        
        current_row = 3
        
        # 1. Success Rate Pie Chart
        current_row = self._add_success_rate_chart(ws, results, current_row)
        current_row += 15
        
        # 2. Metric Performance Bar Chart
        current_row = self._add_metric_performance_chart(ws, results, current_row)
        current_row += 15
        
        # 3. Response Time Timeline (if timing data available)
        timing_stats = results.get_timing_stats()
        if timing_stats['total'] > 0:
            current_row = self._add_response_time_chart(ws, results, current_row)
    
    def _add_metric_summary_chart(self, ws, results: 'EvaluationResult', start_row: int) -> None:
        """Add metric summary bar chart to summary sheet."""
        # Prepare data
        metrics = results.metrics
        means = [results.get_metric_stats(metric)['mean'] for metric in metrics]
        
        if not metrics:
            return
        
        # Create data table for chart
        ws.cell(row=start_row, column=1, value="Metric")
        ws.cell(row=start_row, column=2, value="Mean Score")
        
        for i, (metric, mean) in enumerate(zip(metrics, means)):
            ws.cell(row=start_row + 1 + i, column=1, value=metric)
            ws.cell(row=start_row + 1 + i, column=2, value=mean)
        
        # Create chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Metric Performance Summary"
        chart.y_axis.title = 'Mean Score'
        chart.x_axis.title = 'Metrics'
        
        # Chart data
        data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(metrics))
        cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(metrics))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # Style
        chart.width = 15
        chart.height = 8
        
        # Add to worksheet
        ws.add_chart(chart, f"D{start_row}")
    
    def _add_success_rate_chart(self, ws, results, start_row: int) -> int:  # EvaluationResult object
        """Add success rate pie chart."""
        # Data
        successful = len(results.results)
        failed = len(results.errors)
        
        # Create data table
        ws.cell(row=start_row, column=1, value="Status")
        ws.cell(row=start_row, column=2, value="Count")
        ws.cell(row=start_row + 1, column=1, value="Successful")
        ws.cell(row=start_row + 1, column=2, value=successful)
        ws.cell(row=start_row + 2, column=1, value="Failed")
        ws.cell(row=start_row + 2, column=2, value=failed)
        
        # Create pie chart
        chart = PieChart()
        chart.title = "Success Rate Distribution"
        
        # Chart data
        data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + 2)
        cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + 2)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # Style
        chart.width = 15
        chart.height = 10
        
        # Add to worksheet
        ws.add_chart(chart, f"D{start_row}")
        
        return start_row + 3
    
    def _add_metric_performance_chart(self, ws, results, start_row: int) -> int:  # EvaluationResult object
        """Add metric performance bar chart."""
        metrics = results.metrics
        if not metrics:
            return start_row
        
        # Create data table
        ws.cell(row=start_row, column=1, value="Metric")
        ws.cell(row=start_row, column=2, value="Mean")
        ws.cell(row=start_row, column=3, value="Min")
        ws.cell(row=start_row, column=4, value="Max")
        
        for i, metric in enumerate(metrics):
            stats = results.get_metric_stats(metric)
            ws.cell(row=start_row + 1 + i, column=1, value=metric)
            ws.cell(row=start_row + 1 + i, column=2, value=stats['mean'])
            ws.cell(row=start_row + 1 + i, column=3, value=stats['min'])
            ws.cell(row=start_row + 1 + i, column=4, value=stats['max'])
        
        # Create bar chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Metric Performance Comparison"
        chart.y_axis.title = 'Score'
        chart.x_axis.title = 'Metrics'
        
        # Chart data
        data = Reference(ws, min_col=2, min_row=start_row, max_col=4, max_row=start_row + len(metrics))
        cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(metrics))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # Style
        chart.width = 18
        chart.height = 12
        
        # Add to worksheet
        ws.add_chart(chart, f"F{start_row}")
        
        return start_row + len(metrics) + 1
    
    def _add_response_time_chart(self, ws, results, start_row: int) -> int:  # EvaluationResult object
        """Add response time timeline chart."""
        # Extract timing data
        times = []
        for i, (item_id, result) in enumerate(results.results.items()):
            if 'time' in result and isinstance(result['time'], (int, float)):
                times.append((i + 1, float(result['time'])))
            else:
                times.append((i + 1, 0.0))
        
        if not times:
            return start_row
        
        # Create data table
        ws.cell(row=start_row, column=1, value="Sequence")
        ws.cell(row=start_row, column=2, value="Response Time (s)")
        
        for i, (seq, time_val) in enumerate(times):
            ws.cell(row=start_row + 1 + i, column=1, value=seq)
            ws.cell(row=start_row + 1 + i, column=2, value=time_val)
        
        # Create line chart
        chart = LineChart()
        chart.title = "Response Time Timeline"
        chart.style = 13
        chart.y_axis.title = 'Response Time (seconds)'
        chart.x_axis.title = 'Evaluation Sequence'
        
        # Chart data
        data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(times))
        cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(times))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # Style
        chart.width = 18
        chart.height = 10
        
        # Add to worksheet
        ws.add_chart(chart, f"D{start_row}")
        
        return start_row + len(times) + 1
    
    def create_metric_distribution_chart_data(
        self, 
        results,  # EvaluationResult object
        metric_name: str
    ) -> Tuple[List[float], Dict[str, Any]]:
        """
        Extract metric distribution data for Excel chart creation.
        
        Args:
            results: EvaluationResult object
            metric_name: Name of the metric
            
        Returns:
            Tuple of (values_list, statistics_dict)
        """
        values = []
        for result in results.results.values():
            if 'scores' in result and metric_name in result['scores']:
                score = result['scores'][metric_name]
                if isinstance(score, (int, float)):
                    values.append(float(score))
                elif isinstance(score, bool):
                    values.append(1.0 if score else 0.0)
        
        stats = {}
        if values:
            stats = {
                'mean': statistics.mean(values),
                'std': statistics.stdev(values) if len(values) > 1 else 0.0,
                'min': min(values),
                'max': max(values),
                'count': len(values)
            }
        
        return values, stats