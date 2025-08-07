"""Evaluation results container and analysis."""

from typing import Any, Dict, List, Optional
from datetime import datetime
import statistics
import json
import csv
from pathlib import Path
from rich.console import Console
from rich.table import Table
from rich.panel import Panel

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from openpyxl.chart.label import DataLabelList
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Visualization components will be imported lazily to avoid circular dependencies
VISUALIZATION_AVAILABLE = None


console = Console()


def _check_visualization_available():
    """Check if visualization components are available (lazy loading)."""
    global VISUALIZATION_AVAILABLE
    if VISUALIZATION_AVAILABLE is None:
        try:
            from ..visualizations.charts import ChartGenerator
            from ..visualizations.excel_export import ExcelChartExporter
            from ..visualizations.utils import create_evaluation_report
            VISUALIZATION_AVAILABLE = True
        except ImportError:
            VISUALIZATION_AVAILABLE = False
    return VISUALIZATION_AVAILABLE


class EvaluationResult:
    """Container for evaluation results with analysis capabilities."""
    
    def __init__(self, dataset_name: str, run_name: str, metrics: List[str]):
        """
        Initialize results container.
        
        Args:
            dataset_name: Name of the evaluated dataset
            run_name: Name of this evaluation run
            metrics: List of metric names used
        """
        self.dataset_name = dataset_name
        self.run_name = run_name
        self.metrics = metrics
        self.start_time = datetime.now()
        self.end_time = None
        
        # Results storage
        self.results = {}  # item_id -> result dict
        self.errors = {}   # item_id -> error message
        
    def add_result(self, item_id: str, result: Dict[str, Any]):
        """Add a successful evaluation result."""
        self.results[item_id] = result
    
    def add_error(self, item_id: str, error: str):
        """Add an evaluation error."""
        self.errors[item_id] = error
    
    def finish(self):
        """Mark evaluation as finished."""
        self.end_time = datetime.now()
    
    @property
    def total_items(self) -> int:
        """Total number of evaluated items."""
        return len(self.results) + len(self.errors)
    
    @property
    def success_rate(self) -> float:
        """Percentage of successful evaluations."""
        if self.total_items == 0:
            return 0.0
        return len(self.results) / self.total_items
    
    @property
    def duration(self) -> Optional[float]:
        """Evaluation duration in seconds."""
        if self.end_time:
            return (self.end_time - self.start_time).total_seconds()
        return None
    
    def get_metric_stats(self, metric_name: str) -> Dict[str, float]:
        """
        Get statistics for a specific metric.
        
        Returns dict with: mean, std, min, max, success_rate
        """
        scores = []
        errors = 0
        
        for result in self.results.values():
            if 'scores' in result and metric_name in result['scores']:
                score = result['scores'][metric_name]
                if isinstance(score, (int, float)):
                    scores.append(float(score))
                elif isinstance(score, bool):
                    scores.append(1.0 if score else 0.0)
                elif isinstance(score, dict) and 'error' in score:
                    errors += 1
        
        if not scores:
            return {
                'mean': 0.0,
                'std': 0.0,
                'min': 0.0,
                'max': 0.0,
                'success_rate': 0.0
            }
        
        return {
            'mean': statistics.mean(scores),
            'std': statistics.stdev(scores) if len(scores) > 1 else 0.0,
            'min': min(scores),
            'max': max(scores),
            'success_rate': len(scores) / (len(scores) + errors)
        }
    
    def get_timing_stats(self) -> Dict[str, float]:
        """
        Get timing statistics for all evaluations.
        
        Returns dict with: mean, std, min, max, total
        """
        times = []
        
        for result in self.results.values():
            if 'time' in result and isinstance(result['time'], (int, float)):
                times.append(float(result['time']))
        
        if not times:
            return {
                'mean': 0.0,
                'std': 0.0,
                'min': 0.0,
                'max': 0.0,
                'total': 0.0
            }
        
        return {
            'mean': statistics.mean(times),
            'std': statistics.stdev(times) if len(times) > 1 else 0.0,
            'min': min(times),
            'max': max(times),
            'total': sum(times)
        }
    
    def get_summary(self) -> Dict[str, Any]:
        """
        Get summary statistics as a dictionary.
        
        Returns:
            Dictionary with summary statistics including success_rate,
            average_scores, execution_time, and other metrics.
        """
        # Calculate average scores for each metric
        average_scores = {}
        for metric in self.metrics:
            stats = self.get_metric_stats(metric)
            average_scores[metric] = stats['mean']
        
        return {
            'success_rate': self.success_rate,
            'average_scores': average_scores,
            'execution_time': self.duration or 0.0,
            'total_items': self.total_items,
            'successful_items': len(self.results),
            'failed_items': len(self.errors),
            'metric_stats': {metric: self.get_metric_stats(metric) for metric in self.metrics},
            'timing_stats': self.get_timing_stats()
        }
    
    def summary(self) -> str:
        """Generate a text summary of results."""
        lines = []
        lines.append(f"Evaluation Results: {self.run_name}")
        lines.append(f"Dataset: {self.dataset_name}")
        lines.append(f"Total Items: {self.total_items}")
        lines.append(f"Success Rate: {self.success_rate:.1%}")
        
        if self.duration:
            lines.append(f"Duration: {self.duration:.1f}s")
        
        lines.append("\nMetric Results:")
        for metric in self.metrics:
            stats = self.get_metric_stats(metric)
            lines.append(f"  {metric}:")
            lines.append(f"    Mean: {stats['mean']:.3f}")
            lines.append(f"    Std:  {stats['std']:.3f}")
            lines.append(f"    Range: [{stats['min']:.3f}, {stats['max']:.3f}]")
        
        if self.errors:
            lines.append(f"\nErrors: {len(self.errors)} items failed")
            # Show first few errors
            for item_id, error in list(self.errors.items())[:3]:
                lines.append(f"  - {item_id}: {error[:100]}...")
        
        return "\n".join(lines)
    
    def print_summary(self):
        """Print a rich formatted summary to console."""
        # Create summary table
        table = Table(title=f"Evaluation Results: {self.run_name}")
        table.add_column("Metric", style="cyan")
        table.add_column("Mean", justify="right", style="green")
        table.add_column("Std Dev", justify="right")
        table.add_column("Min", justify="right")
        table.add_column("Max", justify="right")
        table.add_column("Success", justify="right", style="yellow")
        
        for metric in self.metrics:
            stats = self.get_metric_stats(metric)
            table.add_row(
                metric,
                f"{stats['mean']:.3f}",
                f"{stats['std']:.3f}",
                f"{stats['min']:.3f}",
                f"{stats['max']:.3f}",
                f"{stats['success_rate']:.1%}"
            )
        
        # Get timing statistics
        timing_stats = self.get_timing_stats()
        
        # Print overview panel
        overview = Panel(
            f"[bold]Dataset:[/bold] {self.dataset_name}\n"
            f"[bold]Total Items:[/bold] {self.total_items}\n"
            f"[bold]Success Rate:[/bold] {self.success_rate:.1%}\n"
            f"[bold]Total Duration:[/bold] {self.duration:.1f}s\n"
            f"[bold]Average Item Time:[/bold] {timing_stats['mean']:.2f}s ± {timing_stats['std']:.2f}s\n"
            f"[bold]Time Range:[/bold] [{timing_stats['min']:.2f}s, {timing_stats['max']:.2f}s]" if self.duration else "",
            title="Overview",
            expand=False
        )
        
        console.print(overview)
        console.print(table)
        
        if self.errors:
            console.print(f"\n[red]Errors:[/red] {len(self.errors)} items failed")
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert results to dictionary format."""
        metric_stats = {
            metric: self.get_metric_stats(metric) 
            for metric in self.metrics
        }
        
        return {
            'dataset_name': self.dataset_name,
            'run_name': self.run_name,
            'start_time': self.start_time.isoformat(),
            'end_time': self.end_time.isoformat() if self.end_time else None,
            'duration': self.duration,
            'total_items': self.total_items,
            'success_rate': self.success_rate,
            'metrics': self.metrics,
            'metric_stats': metric_stats,
            'results': self.results,
            'errors': self.errors
        }
    
    def failed_items(self) -> List[str]:
        """Get list of failed item IDs."""
        return list(self.errors.keys())
    
    def successful_items(self) -> List[str]:
        """Get list of successful item IDs."""
        return list(self.results.keys())
    
    def save_json(self, filepath: Optional[str] = None) -> str:
        """
        Save results to JSON file.
        
        Args:
            filepath: Optional custom filepath. If not provided, generates one.
            
        Returns:
            Path to the saved file
        """
        if filepath is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = f"eval_results_{self.dataset_name}_{timestamp}.json"
        
        filepath = Path(filepath)
        filepath.parent.mkdir(parents=True, exist_ok=True)
        
        with open(filepath, 'w') as f:
            json.dump(self.to_dict(), f, indent=2, default=str)
        
        console.print(f"[green]✅ Results saved to:[/green] {filepath}")
        return str(filepath)
    
    def save_csv(self, filepath: Optional[str] = None) -> str:
        """
        Save results to CSV file for spreadsheet analysis.
        
        Args:
            filepath: Optional custom filepath. If not provided, generates one.
            
        Returns:
            Path to the saved file
        """
        if filepath is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = f"eval_results_{self.dataset_name}_{timestamp}.csv"
        
        filepath = Path(filepath)
        filepath.parent.mkdir(parents=True, exist_ok=True)
        
        # Prepare rows for CSV
        rows = []
        for item_id, result in self.results.items():
            row = {
                'item_id': item_id,
                'output': str(result.get('output', ''))[:100],  # Truncate long outputs
                'success': result.get('success', False),
                'time': result.get('time', 0.0)
            }
            
            # Add metric scores
            if 'scores' in result:
                for metric_name, score in result['scores'].items():
                    if isinstance(score, dict) and 'error' in score:
                        row[f'metric_{metric_name}'] = 'ERROR'
                        row[f'metric_{metric_name}_error'] = score['error']
                    else:
                        row[f'metric_{metric_name}'] = score
            
            rows.append(row)
        
        # Add failed items
        for item_id, error in self.errors.items():
            row = {
                'item_id': item_id,
                'output': f'ERROR: {error[:100]}',
                'success': False,
                'time': 0.0
            }
            for metric in self.metrics:
                row[f'metric_{metric}'] = 'N/A'
            rows.append(row)
        
        # Write CSV
        if rows:
            with open(filepath, 'w', newline='') as f:
                writer = csv.DictWriter(f, fieldnames=rows[0].keys())
                writer.writeheader()
                writer.writerows(rows)
        
        console.print(f"[green]✅ Results saved to:[/green] {filepath}")
        return str(filepath)
    
    def save_excel(self, filepath: Optional[str] = None) -> str:
        """
        Save results to Excel file with professional formatting and multiple worksheets.
        
        Features:
        - Results worksheet with detailed evaluation data
        - Summary worksheet with aggregated statistics
        - Charts worksheet with visualizations
        - Professional styling and formatting
        
        Args:
            filepath: Optional custom filepath. If not provided, generates one.
            
        Returns:
            Path to the saved file
            
        Raises:
            ImportError: If openpyxl is not installed
        """
        if not EXCEL_AVAILABLE:
            raise ImportError(
                "Excel export requires openpyxl. Install it with: pip install openpyxl"
            )
        
        if filepath is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = f"eval_results_{self.dataset_name}_{timestamp}.xlsx"
        
        filepath = Path(filepath)
        filepath.parent.mkdir(parents=True, exist_ok=True)
        
        # Create workbook with multiple worksheets
        wb = Workbook()
        
        # Remove default sheet and create our custom sheets
        wb.remove(wb.active)
        
        # Create worksheets
        results_ws = wb.create_sheet("Results")
        summary_ws = wb.create_sheet("Summary") 
        charts_ws = wb.create_sheet("Charts")
        
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # 1. RESULTS WORKSHEET
        self._create_results_worksheet(results_ws, header_font, header_fill, border, center_alignment)
        
        # 2. SUMMARY WORKSHEET  
        self._create_summary_worksheet(summary_ws, header_font, header_fill, border, center_alignment)
        
        # 3. CHARTS WORKSHEET
        self._create_charts_worksheet(charts_ws, results_ws, summary_ws)
        
        # Save the workbook
        wb.save(filepath)
        
        console.print(f"[green]Excel report saved to:[/green] {filepath}")
        console.print(f"[blue]Contains:[/blue] Results, Summary, and Charts worksheets")
        return str(filepath)
    
    def _create_results_worksheet(self, ws, header_font, header_fill, border, center_alignment):
        """Create the detailed results worksheet."""
        # Headers
        headers = ['Item ID', 'Output (Preview)', 'Success', 'Time (s)']
        for metric in self.metrics:
            headers.extend([f'{metric} Score', f'{metric} Status'])
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Write data rows
        row = 2
        
        # Successful results
        for item_id, result in self.results.items():
            ws.cell(row=row, column=1, value=item_id).border = border
            
            # Output preview (truncated)
            output_preview = str(result.get('output', ''))[:100]
            if len(str(result.get('output', ''))) > 100:
                output_preview += "..."
            ws.cell(row=row, column=2, value=output_preview).border = border
            
            ws.cell(row=row, column=3, value="Success").border = border
            ws.cell(row=row, column=4, value=result.get('time', 0.0)).border = border
            
            col = 5
            # Add metric scores
            if 'scores' in result:
                for metric in self.metrics:
                    if metric in result['scores']:
                        score = result['scores'][metric]
                        if isinstance(score, dict) and 'error' in score:
                            ws.cell(row=row, column=col, value="ERROR").border = border
                            ws.cell(row=row, column=col+1, value=score['error'][:50]).border = border
                        else:
                            ws.cell(row=row, column=col, value=score).border = border
                            ws.cell(row=row, column=col+1, value="OK").border = border
                    else:
                        ws.cell(row=row, column=col, value="N/A").border = border
                        ws.cell(row=row, column=col+1, value="Missing").border = border
                    col += 2
            
            row += 1
        
        # Failed results
        for item_id, error in self.errors.items():
            ws.cell(row=row, column=1, value=item_id).border = border
            ws.cell(row=row, column=2, value=f"ERROR: {error[:100]}").border = border
            ws.cell(row=row, column=3, value="Failed").border = border
            ws.cell(row=row, column=4, value=0.0).border = border
            
            col = 5
            for metric in self.metrics:
                ws.cell(row=row, column=col, value="N/A").border = border
                ws.cell(row=row, column=col+1, value="Failed").border = border
                col += 2
            
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                    if hasattr(cell, 'value') and cell.value:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 chars
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_summary_worksheet(self, ws, header_font, header_fill, border, center_alignment):
        """Create the summary statistics worksheet."""
        # Overview section
        ws.cell(row=1, column=1, value="EVALUATION OVERVIEW").font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        overview_data = [
            ("Dataset Name", self.dataset_name),
            ("Run Name", self.run_name),
            ("Start Time", self.start_time.strftime("%Y-%m-%d %H:%M:%S")),
            ("End Time", self.end_time.strftime("%Y-%m-%d %H:%M:%S") if self.end_time else "Running"),
            ("Duration", f"{self.duration:.2f}s" if self.duration else "N/A"),
            ("Total Items", self.total_items),
            ("Successful Items", len(self.results)),
            ("Failed Items", len(self.errors)),
            ("Success Rate", f"{self.success_rate:.1%}")
        ]
        
        for i, (label, value) in enumerate(overview_data, 3):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws.cell(row=i, column=2, value=value)
        
        # Metric statistics section
        start_row = len(overview_data) + 5
        ws.cell(row=start_row, column=1, value="METRIC STATISTICS").font = Font(bold=True, size=16)
        ws.merge_cells(f'A{start_row}:F{start_row}')
        
        # Metric stats headers
        metric_headers = ['Metric', 'Mean', 'Std Dev', 'Min', 'Max', 'Success Rate']
        for col, header in enumerate(metric_headers, 1):
            cell = ws.cell(row=start_row+2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Metric stats data
        for i, metric in enumerate(self.metrics, start_row+3):
            stats = self.get_metric_stats(metric)
            ws.cell(row=i, column=1, value=metric).border = border
            ws.cell(row=i, column=2, value=f"{stats['mean']:.3f}").border = border
            ws.cell(row=i, column=3, value=f"{stats['std']:.3f}").border = border
            ws.cell(row=i, column=4, value=f"{stats['min']:.3f}").border = border
            ws.cell(row=i, column=5, value=f"{stats['max']:.3f}").border = border
            ws.cell(row=i, column=6, value=f"{stats['success_rate']:.1%}").border = border
        
        # Timing statistics section
        timing_start = start_row + len(self.metrics) + 5
        ws.cell(row=timing_start, column=1, value="TIMING STATISTICS").font = Font(bold=True, size=16)
        
        timing_stats = self.get_timing_stats()
        timing_data = [
            ("Average Time per Item", f"{timing_stats['mean']:.3f}s"),
            ("Standard Deviation", f"{timing_stats['std']:.3f}s"),
            ("Fastest Item", f"{timing_stats['min']:.3f}s"),
            ("Slowest Item", f"{timing_stats['max']:.3f}s"),
            ("Total Processing Time", f"{timing_stats['total']:.3f}s")
        ]
        
        for i, (label, value) in enumerate(timing_data, timing_start+2):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws.cell(row=i, column=2, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                    if hasattr(cell, 'value') and cell.value:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 40)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_charts_worksheet(self, ws, results_ws, summary_ws):
        """Create charts and visualizations worksheet."""
        # Title
        ws.cell(row=1, column=1, value="EVALUATION CHARTS & VISUALIZATIONS").font = Font(bold=True, size=16)
        ws.merge_cells('A1:H1')
        
        # Chart 1: Success Rate Pie Chart
        if self.total_items > 0:
            pie_chart = PieChart()
            pie_chart.title = "Success vs Failure Rate"
            
            # Create data for pie chart
            ws.cell(row=3, column=1, value="Status")
            ws.cell(row=3, column=2, value="Count")
            ws.cell(row=4, column=1, value="Success")
            ws.cell(row=4, column=2, value=len(self.results))
            ws.cell(row=5, column=1, value="Failed")
            ws.cell(row=5, column=2, value=len(self.errors))
            
            data = Reference(ws, min_col=2, min_row=3, max_row=5)
            labels = Reference(ws, min_col=1, min_row=4, max_row=5)
            pie_chart.add_data(data, titles_from_data=True)
            pie_chart.set_categories(labels)
            
            # Style the pie chart
            pie_chart.dataLabels = DataLabelList()
            pie_chart.dataLabels.showPercent = True
            
            ws.add_chart(pie_chart, "D3")
        
        # Chart 2: Metric Performance Bar Chart
        if self.metrics and len(self.results) > 0:
            bar_chart = BarChart()
            bar_chart.title = "Average Metric Scores"
            bar_chart.y_axis.title = "Score"
            bar_chart.x_axis.title = "Metrics"
            
            # Create data for bar chart  
            start_row = 10
            ws.cell(row=start_row, column=1, value="Metric")
            ws.cell(row=start_row, column=2, value="Average Score")
            
            for i, metric in enumerate(self.metrics, start_row+1):
                stats = self.get_metric_stats(metric)
                ws.cell(row=i, column=1, value=metric)
                ws.cell(row=i, column=2, value=stats['mean'])
            
            data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row+len(self.metrics))
            labels = Reference(ws, min_col=1, min_row=start_row+1, max_row=start_row+len(self.metrics))
            
            bar_chart.add_data(data, titles_from_data=True)
            bar_chart.set_categories(labels)
            
            ws.add_chart(bar_chart, "D12")
        
        # Chart 3: Timing Distribution (if timing data available)
        timing_stats = self.get_timing_stats()
        if timing_stats['total'] > 0:
            # Create a simple timing summary
            ws.cell(row=25, column=1, value="TIMING SUMMARY").font = Font(bold=True, size=14)
            
            timing_data = [
                ("Metric", "Time (seconds)"),
                ("Average", timing_stats['mean']),
                ("Minimum", timing_stats['min']),
                ("Maximum", timing_stats['max']),
                ("Total", timing_stats['total'])
            ]
            
            for i, (label, value) in enumerate(timing_data, 26):
                ws.cell(row=i, column=1, value=label)
                if isinstance(value, (int, float)):
                    ws.cell(row=i, column=2, value=f"{value:.3f}")
                else:
                    ws.cell(row=i, column=2, value=value)
    
    def save(self, format: str = "json", filepath: Optional[str] = None) -> str:
        """
        Save results in specified format.
        
        Args:
            format: Export format - "json", "csv", or "excel"
            filepath: Optional custom filepath
            
        Returns:
            Path to the saved file
        """
        if format.lower() == "json":
            return self.save_json(filepath)
        elif format.lower() == "csv":
            return self.save_csv(filepath)
        elif format.lower() in ["excel", "xlsx"]:
            return self.save_excel(filepath)
        else:
            raise ValueError(f"Unsupported format: {format}. Use 'json', 'csv', or 'excel'.")
    
    # Visualization Methods
    
    def create_dashboard(self, **kwargs):
        """
        Create executive dashboard with multiple charts.
        
        Returns:
            Plotly figure object (if visualization available)
        """
        if not _check_visualization_available():
            raise ImportError("Visualization components not available. Install plotly: pip install plotly")
        
        from ..visualizations.charts import ChartGenerator
        chart_gen = ChartGenerator(**kwargs)
        return chart_gen.create_dashboard(self)
    
    def create_metric_distribution_chart(self, metric_name: str, chart_type: str = 'histogram', **kwargs):
        """
        Create distribution chart for a specific metric.
        
        Args:
            metric_name: Name of the metric to visualize
            chart_type: 'histogram', 'box', or 'violin'
            
        Returns:
            Plotly figure object (if visualization available)
        """
        if not _check_visualization_available():
            raise ImportError("Visualization components not available. Install plotly: pip install plotly")
        
        from ..visualizations.charts import ChartGenerator
        chart_gen = ChartGenerator(**kwargs)
        return chart_gen.create_metric_distribution_chart(self, metric_name, chart_type)
    
    def create_performance_timeline_chart(self, **kwargs):
        """
        Create timeline chart showing response times over evaluation sequence.
        
        Returns:
            Plotly figure object (if visualization available)
        """
        if not _check_visualization_available():
            raise ImportError("Visualization components not available. Install plotly: pip install plotly")
        
        from ..visualizations.charts import ChartGenerator
        chart_gen = ChartGenerator(**kwargs)
        return chart_gen.create_performance_timeline_chart(self)
    
    def create_correlation_chart(self, metric1: str, metric2: str, **kwargs):
        """
        Create scatter plot showing correlation between two metrics.
        
        Args:
            metric1: First metric name
            metric2: Second metric name
            
        Returns:
            Plotly figure object (if visualization available)
        """
        if not _check_visualization_available():
            raise ImportError("Visualization components not available. Install plotly: pip install plotly")
        
        from ..visualizations.charts import ChartGenerator
        chart_gen = ChartGenerator(**kwargs)
        return chart_gen.create_metric_correlation_chart(self, metric1, metric2)
    
    def create_multi_metric_comparison_chart(self, chart_type: str = 'radar', **kwargs):
        """
        Create comparison chart for multiple metrics.
        
        Args:
            chart_type: 'radar', 'bar', or 'line'
            
        Returns:
            Plotly figure object (if visualization available)
        """
        if not _check_visualization_available():
            raise ImportError("Visualization components not available. Install plotly: pip install plotly")
        
        from ..visualizations.charts import ChartGenerator
        chart_gen = ChartGenerator(**kwargs)
        return chart_gen.create_multi_metric_comparison_chart(self, chart_type)
    
    def create_evaluation_report(
        self,
        output_dir: str = "./reports", 
        formats: List[str] = ["html", "excel"],
        include_charts: List[str] = ["dashboard", "distributions", "timeline"]
    ) -> Dict[str, str]:
        """
        Create comprehensive evaluation report with charts.
        
        Args:
            output_dir: Output directory for reports
            formats: List of output formats ("html", "excel", "png", "pdf")
            include_charts: List of chart types to include
            
        Returns:
            Dictionary mapping format to output file path
        """
        if not _check_visualization_available():
            raise ImportError("Visualization components not available. Install plotly: pip install plotly")
        
        from ..visualizations.utils import create_evaluation_report
        return create_evaluation_report(
            self, 
            output_dir=output_dir,
            formats=formats,
            include_charts=include_charts
        )
    
    def export_to_excel_with_charts(
        self, 
        filename: str,
        include_raw_data: bool = True,
        include_charts: bool = True
    ) -> str:
        """
        Export results to Excel with embedded charts.
        
        Args:
            filename: Output Excel filename
            include_raw_data: Whether to include raw evaluation data
            include_charts: Whether to include charts
            
        Returns:
            Path to created Excel file
        """
        if not _check_visualization_available():
            raise ImportError("Visualization components not available. Install openpyxl: pip install openpyxl")
        
        from ..visualizations.excel_export import ExcelChartExporter
        excel_exporter = ExcelChartExporter()
        return excel_exporter.create_excel_report(
            self,
            filename,
            include_raw_data=include_raw_data,
            include_charts=include_charts
        )
    
    def show_dashboard(self, **kwargs):
        """
        Display interactive dashboard in browser/notebook.
        
        This method creates and displays the dashboard chart immediately.
        """
        dashboard = self.create_dashboard(**kwargs)
        dashboard.show()
        return dashboard