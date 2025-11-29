"""Discover and parse historical evaluation runs from local files."""

import csv
import json
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional
import time


@dataclass
class RunInfo:
    """Metadata for a single evaluation run."""
    run_id: str
    task_name: str
    model_name: str
    dataset_name: str
    timestamp: datetime
    file_path: str
    metrics: List[str]
    total_items: int
    success_count: int
    error_count: int
    metric_averages: Dict[str, float] = field(default_factory=dict)
    avg_latency_ms: float = 0.0

    @property
    def success_rate(self) -> float:
        if self.total_items == 0:
            return 0.0
        return self.success_count / self.total_items

    def to_dict(self) -> Dict[str, Any]:
        return {
            "run_id": self.run_id,
            "task_name": self.task_name,
            "model_name": self.model_name,
            "dataset_name": self.dataset_name,
            "timestamp": self.timestamp.isoformat(),
            "file_path": self.file_path,
            "metrics": self.metrics,
            "metric_averages": self.metric_averages,
            "total_items": self.total_items,
            "success_count": self.success_count,
            "error_count": self.error_count,
            "success_rate": self.success_rate,
            "avg_latency_ms": self.avg_latency_ms,
        }


@dataclass
class RunIndex:
    """Hierarchical index: task -> model -> runs."""
    tasks: Dict[str, Dict[str, List[RunInfo]]] = field(default_factory=dict)
    last_updated: datetime = field(default_factory=datetime.now)

    def to_dict(self) -> Dict[str, Any]:
        result = {"tasks": {}, "last_updated": self.last_updated.isoformat()}
        for task_name, models in self.tasks.items():
            result["tasks"][task_name] = {}
            for model_name, runs in models.items():
                result["tasks"][task_name][model_name] = [r.to_dict() for r in runs]
        return result


def strip_model_provider(model_name: str) -> str:
    """Strip provider prefix from model name (e.g., 'qwen/qwen3-coder' -> 'qwen3-coder')."""
    if not model_name:
        return model_name
    parts = model_name.split("/")
    return parts[-1] if len(parts) > 1 else model_name


class RunDiscovery:
    """Scan results directory and build index of historical runs."""

    def __init__(self, results_dir: str = "eval_results"):
        self.results_dir = Path(results_dir)
        self._cache: Optional[RunIndex] = None
        self._cache_time: Optional[float] = None
        self._cache_ttl = 30  # seconds

    def scan(self, force_refresh: bool = False) -> RunIndex:
        """Scan results directory and return hierarchical index."""
        # Check cache
        if not force_refresh and self._cache is not None:
            if self._cache_time and (time.time() - self._cache_time) < self._cache_ttl:
                return self._cache

        index = RunIndex()

        if not self.results_dir.exists():
            self._cache = index
            self._cache_time = time.time()
            return index

        # Traverse: results_dir/{task}/{model}/{date}/*.csv
        for task_dir in self.results_dir.iterdir():
            if not task_dir.is_dir():
                continue
            task_name = task_dir.name

            for model_dir in task_dir.iterdir():
                if not model_dir.is_dir():
                    continue
                model_name = model_dir.name

                for date_dir in model_dir.iterdir():
                    if not date_dir.is_dir():
                        continue

                    # Support both .csv and .xlsx files
                    for result_file in list(date_dir.glob("*.csv")) + list(date_dir.glob("*.xlsx")):
                        run_info = self._parse_result_file(result_file, task_name, model_name)
                        if run_info:
                            if task_name not in index.tasks:
                                index.tasks[task_name] = {}
                            if model_name not in index.tasks[task_name]:
                                index.tasks[task_name][model_name] = []
                            index.tasks[task_name][model_name].append(run_info)

        # Sort runs by timestamp (newest first)
        for task_name in index.tasks:
            for model_name in index.tasks[task_name]:
                index.tasks[task_name][model_name].sort(
                    key=lambda r: r.timestamp, reverse=True
                )

        index.last_updated = datetime.now()
        self._cache = index
        self._cache_time = time.time()
        return index

    def _parse_result_file(
        self, file_path: Path, task_name: str, model_name: str
    ) -> Optional[RunInfo]:
        """Parse result file (CSV or XLSX) to extract run metadata."""
        if file_path.suffix.lower() == ".xlsx":
            return self._parse_xlsx_file(file_path, task_name, model_name)
        return self._parse_csv_file(file_path, task_name, model_name)

    def _parse_csv_file(
        self, file_path: Path, task_name: str, model_name: str
    ) -> Optional[RunInfo]:
        """Parse CSV file to extract run metadata."""
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                fieldnames = reader.fieldnames or []

                # Extract metrics from column names ending with _score
                metrics = [
                    col.replace("_score", "")
                    for col in fieldnames
                    if col.endswith("_score")
                ]

                # Read all rows to count totals
                rows = list(reader)
                total_items = len(rows)

                if total_items == 0:
                    return None

                # Get run info from first row
                first_row = rows[0]
                run_name = first_row.get("run_name", "")
                dataset_name = first_row.get("dataset_name", "unknown")

                # Try to get model from run_metadata JSON and strip provider prefix
                actual_model = strip_model_provider(model_name)
                try:
                    metadata = json.loads(first_row.get("run_metadata", "{}"))
                    if "model" in metadata:
                        actual_model = strip_model_provider(metadata["model"])
                except (json.JSONDecodeError, TypeError):
                    pass

                # Count successes vs errors and calculate metric averages
                error_count = 0
                metric_sums: Dict[str, float] = {m: 0.0 for m in metrics}
                metric_counts: Dict[str, int] = {m: 0 for m in metrics}
                latency_sum = 0.0
                latency_count = 0

                for row in rows:
                    output = row.get("output", "")
                    if output.startswith("ERROR:") or "ERROR" in str(row.get(f"{metrics[0]}_score", "")) if metrics else False:
                        error_count += 1

                    # Accumulate metric scores
                    for m in metrics:
                        score_str = row.get(f"{m}_score", "")
                        try:
                            score = float(score_str)
                            metric_sums[m] += score
                            metric_counts[m] += 1
                        except (ValueError, TypeError):
                            pass

                    # Accumulate latency (time column is in seconds)
                    time_str = row.get("time", "")
                    try:
                        latency_sec = float(time_str)
                        latency_sum += latency_sec * 1000  # Convert to ms
                        latency_count += 1
                    except (ValueError, TypeError):
                        pass

                success_count = total_items - error_count

                # Calculate metric averages
                metric_averages = {}
                for m in metrics:
                    if metric_counts[m] > 0:
                        metric_averages[m] = metric_sums[m] / metric_counts[m]
                    else:
                        metric_averages[m] = 0.0

                # Calculate average latency
                avg_latency_ms = latency_sum / latency_count if latency_count > 0 else 0.0

                # Parse timestamp from run_name
                timestamp = self._extract_timestamp(run_name)

                return RunInfo(
                    run_id=run_name or file_path.stem,
                    task_name=task_name,
                    model_name=actual_model,
                    dataset_name=dataset_name,
                    timestamp=timestamp,
                    file_path=str(file_path),
                    metrics=metrics,
                    total_items=total_items,
                    success_count=success_count,
                    error_count=error_count,
                    metric_averages=metric_averages,
                    avg_latency_ms=avg_latency_ms,
                )
        except Exception:
            return None

    def _parse_xlsx_file(
        self, file_path: Path, task_name: str, model_name: str
    ) -> Optional[RunInfo]:
        """Parse XLSX file to extract run metadata."""
        try:
            import openpyxl
        except ImportError:
            return None

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active

            # Get header from first row
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            fieldnames = [str(cell) if cell else "" for cell in header_row]

            # Read all data rows
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {fieldnames[i]: row[i] for i in range(len(fieldnames)) if i < len(row)}
                rows.append(row_dict)

            wb.close()

            if not rows:
                return None

            # Extract metrics from column names ending with _score
            metrics = [
                col.replace("_score", "")
                for col in fieldnames
                if col.endswith("_score")
            ]

            total_items = len(rows)
            first_row = rows[0]
            run_name = str(first_row.get("run_name", "") or "")
            dataset_name = str(first_row.get("dataset_name", "unknown") or "unknown")

            # Try to get model from run_metadata JSON
            actual_model = strip_model_provider(model_name)
            try:
                metadata_str = first_row.get("run_metadata", "{}")
                if metadata_str:
                    metadata = json.loads(str(metadata_str))
                    if "model" in metadata:
                        actual_model = strip_model_provider(metadata["model"])
            except (json.JSONDecodeError, TypeError):
                pass

            # Count successes vs errors and calculate metric averages
            error_count = 0
            metric_sums: Dict[str, float] = {m: 0.0 for m in metrics}
            metric_counts: Dict[str, int] = {m: 0 for m in metrics}
            latency_sum = 0.0
            latency_count = 0

            for row in rows:
                output = str(row.get("output", "") or "")
                if output.startswith("ERROR:") or (metrics and "ERROR" in str(row.get(f"{metrics[0]}_score", "") or "")):
                    error_count += 1

                # Accumulate metric scores
                for m in metrics:
                    score_val = row.get(f"{m}_score")
                    try:
                        score = float(score_val) if score_val is not None else None
                        if score is not None:
                            metric_sums[m] += score
                            metric_counts[m] += 1
                    except (ValueError, TypeError):
                        pass

                # Accumulate latency
                time_val = row.get("time")
                try:
                    latency_sec = float(time_val) if time_val is not None else None
                    if latency_sec is not None:
                        latency_sum += latency_sec * 1000
                        latency_count += 1
                except (ValueError, TypeError):
                    pass

            success_count = total_items - error_count

            # Calculate metric averages
            metric_averages = {}
            for m in metrics:
                if metric_counts[m] > 0:
                    metric_averages[m] = metric_sums[m] / metric_counts[m]
                else:
                    metric_averages[m] = 0.0

            avg_latency_ms = latency_sum / latency_count if latency_count > 0 else 0.0
            timestamp = self._extract_timestamp(run_name)

            return RunInfo(
                run_id=run_name or file_path.stem,
                task_name=task_name,
                model_name=actual_model,
                dataset_name=dataset_name,
                timestamp=timestamp,
                file_path=str(file_path),
                metrics=metrics,
                total_items=total_items,
                success_count=success_count,
                error_count=error_count,
                metric_averages=metric_averages,
                avg_latency_ms=avg_latency_ms,
            )
        except Exception:
            return None

    def _extract_timestamp(self, run_name: str) -> datetime:
        """Extract timestamp from run_name pattern: task-model-YYMMDD-HHMM[-counter]."""
        match = re.search(r"-(\d{6}-\d{4})(?:-\d+)?$", run_name or "")
        if match:
            try:
                return datetime.strptime(match.group(1), "%y%m%d-%H%M")
            except ValueError:
                pass
        return datetime.now()

    def get_run_data(self, file_path: str) -> Dict[str, Any]:
        """Load full run data and transform to UI snapshot format."""
        path = Path(file_path)
        if not path.exists():
            return {"error": "File not found"}

        # Handle xlsx files
        if path.suffix.lower() == ".xlsx":
            return self._get_xlsx_run_data(path)

        try:
            with open(path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                fieldnames = reader.fieldnames or []
                rows = list(reader)
        except Exception as e:
            return {"error": str(e)}

        if not rows:
            return {"error": "Empty file"}

        # Extract metrics
        metric_names = [
            col.replace("_score", "")
            for col in fieldnames
            if col.endswith("_score")
        ]

        # Build run info
        first_row = rows[0]
        run_name = first_row.get("run_name", "")
        dataset_name = first_row.get("dataset_name", "")

        # Parse config
        config = {}
        try:
            config = json.loads(first_row.get("run_config", "{}"))
        except (json.JSONDecodeError, TypeError):
            pass

        # Parse metadata
        metadata = {}
        try:
            metadata = json.loads(first_row.get("run_metadata", "{}"))
        except (json.JSONDecodeError, TypeError):
            pass

        # Transform rows to UI snapshot format
        ui_rows = []
        stats = {"total": len(rows), "completed": 0, "in_progress": 0, "pending": 0, "failed": 0}

        for idx, row in enumerate(rows):
            output = row.get("output", "")
            is_error = output.startswith("ERROR:") or output.startswith("ERROR ")

            # Determine status
            status = "completed"
            if is_error:
                status = "error"
                stats["failed"] += 1
            else:
                stats["completed"] += 1

            # Parse time
            time_val = row.get("time", "0")
            try:
                latency_ms = float(time_val) * 1000
            except (ValueError, TypeError):
                latency_ms = 0

            # Extract metric values
            metric_values = []
            metric_meta = {}
            for m in metric_names:
                score = row.get(f"{m}_score", "")
                metric_values.append(score)
                # Collect additional metric fields
                meta = {}
                for col in fieldnames:
                    if col.startswith(f"{m}_") and col != f"{m}_score":
                        field_name = col[len(f"{m}_"):]
                        meta[field_name] = row.get(col, "")
                if meta:
                    metric_meta[m] = meta

            # Build UI row
            input_full = row.get("input", "")
            output_full = row.get("output", "")
            expected_full = row.get("expected_output", "")

            ui_row = {
                "index": idx,
                "status": status,
                "input": input_full[:100] if len(input_full) > 100 else input_full,
                "input_full": input_full,
                "output": output_full[:100] if len(output_full) > 100 else output_full,
                "output_full": output_full,
                "expected": expected_full[:100] if len(expected_full) > 100 else expected_full,
                "expected_full": expected_full,
                "time": time_val,
                "latency_ms": latency_ms,
                "trace_id": row.get("trace_id", ""),
                "trace_url": "",  # Historical runs don't have live trace URLs
                "metric_values": metric_values,
                "metric_meta": metric_meta,
            }
            ui_rows.append(ui_row)

        # Calculate success rate
        stats["success_rate"] = (stats["completed"] / stats["total"] * 100) if stats["total"] > 0 else 0

        return {
            "run": {
                "dataset_name": dataset_name,
                "run_name": run_name,
                "metric_names": metric_names,
                "config": config,
                "metadata": metadata,
                "langfuse_host": "",
                "langfuse_project_id": "",
            },
            "snapshot": {
                "rows": ui_rows,
                "stats": stats,
                "metric_names": metric_names,
            },
        }

    def _get_xlsx_run_data(self, path: Path) -> Dict[str, Any]:
        """Load xlsx run data and transform to UI snapshot format."""
        try:
            import openpyxl
        except ImportError:
            return {"error": "openpyxl is required to read xlsx files"}

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active

            # Get header from first row
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            fieldnames = [str(cell) if cell else "" for cell in header_row]

            # Read all data rows
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {fieldnames[i]: row[i] for i in range(len(fieldnames)) if i < len(row)}
                rows.append(row_dict)

            wb.close()
        except Exception as e:
            return {"error": str(e)}

        if not rows:
            return {"error": "Empty file"}

        # Extract metrics
        metric_names = [
            col.replace("_score", "")
            for col in fieldnames
            if col.endswith("_score")
        ]

        # Build run info
        first_row = rows[0]
        run_name = str(first_row.get("run_name", "") or "")
        dataset_name = str(first_row.get("dataset_name", "") or "")

        # Parse config
        config = {}
        try:
            config_str = first_row.get("run_config", "{}")
            if config_str:
                config = json.loads(str(config_str))
        except (json.JSONDecodeError, TypeError):
            pass

        # Parse metadata
        metadata = {}
        try:
            metadata_str = first_row.get("run_metadata", "{}")
            if metadata_str:
                metadata = json.loads(str(metadata_str))
        except (json.JSONDecodeError, TypeError):
            pass

        # Transform rows to UI snapshot format
        ui_rows = []
        stats = {"total": len(rows), "completed": 0, "in_progress": 0, "pending": 0, "failed": 0}

        for idx, row in enumerate(rows):
            output = str(row.get("output", "") or "")
            is_error = output.startswith("ERROR:") or output.startswith("ERROR ")

            # Determine status
            status = "completed"
            if is_error:
                status = "error"
                stats["failed"] += 1
            else:
                stats["completed"] += 1

            # Parse time
            time_val = row.get("time", 0)
            try:
                latency_ms = float(time_val) * 1000 if time_val else 0
            except (ValueError, TypeError):
                latency_ms = 0

            # Extract metric values
            metric_values = []
            metric_meta = {}
            for m in metric_names:
                score = row.get(f"{m}_score", "")
                metric_values.append(score if score is not None else "")
                # Collect additional metric fields
                meta = {}
                for col in fieldnames:
                    if col.startswith(f"{m}_") and col != f"{m}_score":
                        field_name = col[len(f"{m}_"):]
                        meta[field_name] = row.get(col, "")
                if meta:
                    metric_meta[m] = meta

            # Build UI row
            input_full = str(row.get("input", "") or "")
            output_full = str(row.get("output", "") or "")
            expected_full = str(row.get("expected_output", "") or "")

            ui_row = {
                "index": idx,
                "status": status,
                "input": input_full[:100] if len(input_full) > 100 else input_full,
                "input_full": input_full,
                "output": output_full[:100] if len(output_full) > 100 else output_full,
                "output_full": output_full,
                "expected": expected_full[:100] if len(expected_full) > 100 else expected_full,
                "expected_full": expected_full,
                "time": str(time_val) if time_val else "0",
                "latency_ms": latency_ms,
                "trace_id": str(row.get("trace_id", "") or ""),
                "trace_url": "",
                "metric_values": metric_values,
                "metric_meta": metric_meta,
            }
            ui_rows.append(ui_row)

        # Calculate success rate
        stats["success_rate"] = (stats["completed"] / stats["total"] * 100) if stats["total"] > 0 else 0

        return {
            "run": {
                "dataset_name": dataset_name,
                "run_name": run_name,
                "metric_names": metric_names,
                "config": config,
                "metadata": metadata,
                "langfuse_host": "",
                "langfuse_project_id": "",
            },
            "snapshot": {
                "rows": ui_rows,
                "stats": stats,
                "metric_names": metric_names,
            },
        }
